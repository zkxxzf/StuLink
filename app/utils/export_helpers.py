"""学生列表导出工具"""
import io
from datetime import datetime
from flask import send_file
from app.models import Student, Room, BedAssignment, StudentAccommodation, OperationLog
from app.utils.helpers import get_graduated_grades
from app.extensions import db

BASE_COLUMNS = [
    ('student_number', '学号'),
    ('name', '姓名'),
    ('gender', '性别'),
    ('grade', '年级'),
    ('class_name', '班级'),
    ('subject_selection', '选科'),
    ('ethnicity', '民族'),
]

SENSITIVE_COLUMNS = {
    'id_card_number': {
        'label': '身份证号',
        'perm': 'students.export_id_card',
        'width': 18,
    },
    'phone1': {
        'label': '联系方式1',
        'perm': 'students.export_phone',
        'width': 15,
    },
    'phone2': {
        'label': '联系方式2',
        'perm': 'students.export_phone',
        'width': 15,
    },
    'graduation_school': {
        'label': '毕业学校',
        'perm': 'students.export_graduation_school',
        'width': 25,
    },
    'graduation_school_code': {
        'label': '毕业学校代码',
        'perm': 'students.export_graduation_school',
        'width': 12,
    },
    'enrollment_status': {
        'label': '学籍情况',
        'perm': 'students.export_enrollment',
        'width': 12,
    },
    'enrollment_notes': {
        'label': '学籍备注',
        'perm': 'students.export_enrollment',
        'width': 20,
    },
}


def do_export_students(args):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    from flask_login import current_user

    q = Student.query
    gds = get_graduated_grades()
    if gds:
        q = q.filter(~Student.grade.in_(gds))
    if current_user.role == 'homeroom_teacher':
        q = q.filter_by(grade=current_user.grade, class_name=current_user.class_name)
    for k in ['gender', 'grade', 'class_name', 'subject_selection', 'enrollment_status']:
        v = args.get(k)
        if v:
            q = q.filter_by(**{k: v})
    sch = args.get('graduation_school', '').strip()
    if sch:
        q = q.filter(Student.graduation_school.contains(sch))
    name = args.get('name', '').strip()
    if name:
        q = q.filter(Student.name.contains(name))
    student_number = args.get('student_number', '').strip()
    if student_number:
        q = q.filter(Student.student_number.contains(student_number))

    students = q.order_by(Student.grade, Student.class_name, Student.student_number).all()

    requested_columns = args.getlist('columns')

    selected_columns = []
    selected_widths = []

    base_columns_dict = dict(BASE_COLUMNS)

    selected_columns.append(('student_number', base_columns_dict['student_number']))
    selected_widths.append(10)

    for field in requested_columns:
        if field in base_columns_dict and field != 'student_number':
            selected_columns.append((field, base_columns_dict[field]))
            selected_widths.append(10)
        elif field in SENSITIVE_COLUMNS:
            config = SENSITIVE_COLUMNS[field]
            if current_user.has_perm(config['perm']):
                selected_columns.append((field, config['label']))
                selected_widths.append(config['width'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '学生列表'
    hf = Font(bold=True, color='FFFFFF')
    hfl = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    for ci, (field, label) in enumerate(selected_columns, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = hf
        c.fill = hfl
        c.alignment = Alignment(horizontal='center')
        c.border = tb

    for ri, s in enumerate(students, 2):
        row_data = []
        for field, _ in selected_columns:
            if field == 'id_card_number':
                value = s.id_card_number or ''
            else:
                value = getattr(s, field, '') or ''
            row_data.append(value)
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = tb

    for i, w in enumerate(selected_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    download_name = f'学生列表_{timestamp}.xlsx'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    try:
        import json
        log_detail = {
            'columns': [label for _, label in selected_columns],
            'record_count': len(students),
            'file_name': download_name,
            'filters': {
                'gender': args.get('gender'),
                'grade': args.get('grade'),
                'class_name': args.get('class_name'),
                'name': args.get('name'),
                'student_number': args.get('student_number'),
            }
        }
        log = OperationLog(
            user_id=current_user.id,
            action='导出',
            target_type='学生',
            module='system',
            detail=json.dumps(log_detail, ensure_ascii=False),
            ip_address=args.get('ip_address', '')
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        db.session.rollback()

    return send_file(out, as_attachment=True, download_name=download_name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


ACCOMMODATION_BASE_COLUMNS = [
    ('student_number', '学号'),
    ('name', '姓名'),
    ('gender', '性别'),
    ('grade', '年级'),
    ('class_name', '班级'),
    ('subject_selection', '选科'),
    ('boarding_type', '住校/走读'),
    ('day_student_type', '出门权限'),
    ('room', '宿舍'),
    ('bed', '床位'),
]

ACCOMMODATION_SENSITIVE_COLUMNS = {
    'phone1': {
        'label': '联系方式1',
        'perm': 'students.export_phone',
        'width': 15,
    },
    'phone2': {
        'label': '联系方式2',
        'perm': 'students.export_phone',
        'width': 15,
    },
}


def do_export_student_accommodation(args):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    from flask_login import current_user

    q = Student.query
    gds = get_graduated_grades()
    if gds:
        q = q.filter(~Student.grade.in_(gds))

    for k in ['gender', 'grade', 'class_name', 'subject_selection']:
        v = args.get(k)
        if v:
            q = q.filter_by(**{k: v})

    bt = args.get('boarding_type')
    if bt:
        acc_ids = [sa.student_id for sa in StudentAccommodation.query.filter_by(boarding_type=bt).all()]
        if acc_ids:
            q = q.filter(Student.id.in_(acc_ids))
        else:
            q = q.filter(Student.id == -1)

    dst = args.get('day_student_type')
    if dst:
        acc_ids = [sa.student_id for sa in StudentAccommodation.query.filter_by(day_student_type=dst).all()]
        if acc_ids:
            q = q.filter(Student.id.in_(acc_ids))
        else:
            q = q.filter(Student.id == -1)

    name = args.get('name', '').strip()
    if name:
        q = q.filter(Student.name.contains(name))
    student_number = args.get('student_number', '').strip()
    if student_number:
        q = q.filter(Student.student_number.contains(student_number))
    room_number = args.get('room_number', '').strip()
    if room_number:
        bed_sub = db.session.query(BedAssignment.student_id).join(BedAssignment.room).filter(
            Room.room_number.contains(room_number)
        ).filter(BedAssignment.student_id.isnot(None)).all()
        bed_ids = [b[0] for b in bed_sub if b[0]]
        if bed_ids:
            q = q.filter(Student.id.in_(bed_ids))
        else:
            q = q.filter(Student.id == -1)

    students = q.order_by(Student.grade, Student.class_name, Student.student_number).all()

    requested_columns = args.getlist('columns')

    selected_columns = []
    selected_widths = []

    base_columns_dict = dict(ACCOMMODATION_BASE_COLUMNS)

    selected_columns.append(('student_number', base_columns_dict['student_number']))
    selected_widths.append(10)

    for field in requested_columns:
        if field in base_columns_dict and field != 'student_number':
            selected_columns.append((field, base_columns_dict[field]))
            selected_widths.append(10)
        elif field in ACCOMMODATION_SENSITIVE_COLUMNS:
            config = ACCOMMODATION_SENSITIVE_COLUMNS[field]
            if current_user.has_perm(config['perm']):
                selected_columns.append((field, config['label']))
                selected_widths.append(config['width'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '学生住宿'
    hf = Font(bold=True, color='FFFFFF')
    hfl = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    for ci, (field, label) in enumerate(selected_columns, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = hf
        c.fill = hfl
        c.alignment = Alignment(horizontal='center')
        c.border = tb

    for ri, s in enumerate(students, 2):
        row_data = []
        room_info = '-'
        bed_info = '-'
        if s.bed_assignment and s.bed_assignment.room:
            room_info = f"{s.bed_assignment.room.building} {s.bed_assignment.room.room_number}"
            bed_info = f"{s.bed_assignment.bed_number}床"
        acc = s.accommodation

        for field, _ in selected_columns:
            if field == 'room':
                value = room_info
            elif field == 'bed':
                value = bed_info
            elif field == 'boarding_type':
                value = acc.boarding_type if acc else '-'
            elif field == 'day_student_type':
                value = acc.day_student_type if acc else '-'
            elif field == 'phone1':
                value = s.phone1 or ''
            elif field == 'phone2':
                value = s.phone2 or ''
            else:
                value = getattr(s, field, '') or ''
            row_data.append(value)

        for ci, v in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = tb

    for i, w in enumerate(selected_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    download_name = f'学生住宿_{timestamp}.xlsx'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    try:
        import json
        log_detail = {
            'columns': [label for _, label in selected_columns],
            'record_count': len(students),
            'file_name': download_name,
            'filters': {
                'gender': args.get('gender'),
                'grade': args.get('grade'),
                'class_name': args.get('class_name'),
                'name': args.get('name'),
                'student_number': args.get('student_number'),
                'boarding_type': args.get('boarding_type'),
                'day_student_type': args.get('day_student_type'),
                'room_number': args.get('room_number'),
            }
        }
        log = OperationLog(
            user_id=current_user.id,
            action='导出',
            target_type='学生住宿',
            module='dormitory',
            detail=json.dumps(log_detail, ensure_ascii=False),
            ip_address=args.get('ip_address', '')
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        db.session.rollback()

    return send_file(out, as_attachment=True, download_name=download_name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
