# StuLink v1.7.0 2026-08-02
# Copyright (c) 2026 zkxxzf. Apache License 2.0
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from flask_login import login_required, current_user
from app.extensions import db
from app.models import Room, BedAssignment, Student, StudentAccommodation
from app.utils.decorators import perm_required
from app.utils.helpers import get_dict_values, log_operation
from sqlalchemy import func
import json
import re

bp = Blueprint('rooms', __name__, url_prefix='/rooms')


def _class_sort_key(name):
    """合班名班号排序：按数字升序（01班 < 02班 < 10班），无数字按文本"""
    m = re.search(r'(\d+)', name)
    return (0, int(m.group(1))) if m else (1, name)


@bp.route('/')
@login_required
def list_rooms():
    query = Room.query.filter_by(is_active=True)

    building = request.args.get('building', '')
    gender = request.args.get('gender', '')
    floor = request.args.get('floor', '')
    capacity = request.args.get('capacity', '')

    if building:
        query = query.filter_by(building=building)
    if gender:
        query = query.filter_by(gender=gender)
    if floor:
        # 去掉可能的"层"字，提取数字
        floor_num = floor.replace('层', '').strip()
        try:
            query = query.filter_by(floor=int(floor_num))
        except ValueError:
            pass  # 如果转换失败，忽略该筛选条件
    if capacity:
        try:
            query = query.filter_by(capacity=int(capacity))
        except ValueError:
            pass

    rooms = query.order_by(Room.building, Room.floor, Room.room_number).all()

    # 优化：使用单次聚合查询获取所有房间的入住人数，避免 N+1 查询
    occupancy_data = db.session.query(
        BedAssignment.room_id,
        func.count(BedAssignment.id).label('count')
    ).filter(
        BedAssignment.student_id.isnot(None)
    ).group_by(BedAssignment.room_id).all()

    # 转换为字典便于查找
    occupancy_map = {row.room_id: row.count for row in occupancy_data}

    # 批量查询每间房的学生姓名（跨库分两步查询，避免 N+1）
    room_ids = [r.id for r in rooms]
    students_by_room = {}
    if room_ids:
        # 第一步：查询每间房的床位分配（dormitory.db）
        bed_data = BedAssignment.query.filter(
            BedAssignment.room_id.in_(room_ids),
            BedAssignment.student_id.isnot(None)
        ).order_by(BedAssignment.room_id, BedAssignment.bed_number).all()
        # 第二步：批量查询学生姓名（system.db）
        student_ids = list(set(b.student_id for b in bed_data))
        student_map = {}
        if student_ids:
            students = Student.query.filter(Student.id.in_(student_ids)).all()
            student_map = {s.id: s.name for s in students}
        # 按房间分组
        for b in bed_data:
            name = student_map.get(b.student_id, '?')
            students_by_room.setdefault(b.room_id, []).append(name)

    room_data = []
    for room in rooms:
        # 构建班级显示信息（合班自动识别，兼容历史数据：合班名可能存于 combined_class）
        cn = room.combined_name or room.class_name
        class_info = None
        if room.grade and cn:
            class_info = f"{room.grade} {cn}"
        elif room.grade:
            class_info = room.grade
        elif cn:
            class_info = cn

        room_data.append({
            'room': room,
            'occupancy': occupancy_map.get(room.id, 0),
            'class_info': class_info,
            'student_names': students_by_room.get(room.id, []),
        })

    return render_template('dormitory/rooms/list.html', room_data=room_data,
                           buildings=get_dict_values('building'),
                           floors=get_dict_values('floor'))


@bp.route('/<int:id>')
@login_required
def detail(id):
    room = Room.query.get_or_404(id)
    beds = BedAssignment.query.filter_by(room_id=room.id).order_by(
        BedAssignment.bed_number).all()
    
    # 优化：批量获取所有学生信息，避免 N+1 查询
    student_ids = [bed.student_id for bed in beds if bed.student_id]
    if student_ids:
        students = Student.query.filter(Student.id.in_(student_ids)).all()
        student_map = {s.id: s for s in students}
    else:
        student_map = {}
    
    for bed in beds:
        bed.student_info = student_map.get(bed.student_id) if bed.student_id else None
    
    return render_template('dormitory/rooms/detail.html', room=room, beds=beds)


@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@perm_required('dormitory.manage')
def edit(id):
    room = Room.query.get_or_404(id)
    if request.method == 'POST':
        new_gender = request.form.get('gender', room.gender)
        new_capacity = int(request.form.get('capacity', room.capacity))
        new_grade = request.form.get('grade', '') or None
        new_class = request.form.get('class_name', '') or None
        new_notes = request.form.get('notes', '') or None

        # 检查房间是否已分配班级或已安排床铺
        has_assignment = room.class_name and room.class_name.strip()
        bed_count = BedAssignment.query.filter(
            BedAssignment.room_id == room.id,
            BedAssignment.student_id.isnot(None)
        ).count()
        has_beds = bed_count > 0

        gender_changed = new_gender != room.gender
        capacity_changed = new_capacity != room.capacity

        if (has_assignment or has_beds) and (gender_changed or capacity_changed):
            reasons = []
            if has_assignment:
                reasons.append(f'已分配给班级"{room.class_name}"')
            if has_beds:
                reasons.append(f'已有{bed_count}名学生入住')
            flash(f'该宿舍{",".join(reasons)}，请先清除宿舍分配和床位后再修改性别或床位数', 'danger')
            return redirect(url_for('rooms.detail', id=room.id))

        old_capacity = room.capacity
        if new_capacity != old_capacity:
            _adjust_beds(room, old_capacity, new_capacity)

        room.gender = new_gender
        room.capacity = new_capacity
        room.grade = new_grade
        room.class_name = new_class
        # 自动识别合班：class_name 含多个班（+）即视为合班宿舍，无需手动设置合班标记
        # 手动编辑不维护合班详情（自动分配才会写入各班级人数份额）
        room.combined_details = None
        if new_class and '+' in new_class:
            room.combined_class = new_class
        else:
            room.combined_class = None
        room.notes = new_notes
        db.session.commit()
        flash(f'{room.display_name} 信息已更新', 'success')
        return redirect(url_for('rooms.detail', id=room.id))

    return render_template('dormitory/rooms/form.html', room=room, title='编辑宿舍',
                           grades=get_dict_values('grade'), classes=get_dict_values('class'),
                           buildings=get_dict_values('building'), floors=get_dict_values('floor'))


@bp.route('/create', methods=['GET', 'POST'])
@perm_required('dormitory.manage')
def create():
    if request.method == 'POST':
        building = request.form.get('building', '').strip()
        room_number = request.form.get('room_number', '').strip()
        floor = request.form.get('floor', '').strip()
        gender = request.form.get('gender', '男')
        capacity = int(request.form.get('capacity', 8))
        notes = request.form.get('notes', '') or None

        if not room_number or not building:
            flash('请输入宿舍楼和房间号', 'danger')
            return render_template('dormitory/rooms/form.html', room=None, title='新增宿舍',
                                   grades=get_dict_values('grade'), classes=get_dict_values('class'),
                                   buildings=get_dict_values('building'), floors=get_dict_values('floor'))

        if Room.query.filter_by(building=building, room_number=room_number).first():
            flash(f'{building} {room_number} 已存在', 'danger')
            return render_template('dormitory/rooms/form.html', room=None, title='新增宿舍',
                                   grades=get_dict_values('grade'), classes=get_dict_values('class'),
                                   buildings=get_dict_values('building'), floors=get_dict_values('floor'))

        # 优先使用选择的楼层，如果没有则从房间号提取
        floor_num = 1
        if floor:
            try:
                floor_num = int(floor)
            except ValueError:
                for ch in room_number:
                    if ch.isdigit():
                        floor_num = int(ch)
                        break
        else:
            for ch in room_number:
                if ch.isdigit():
                    floor_num = int(ch)
                    break

        room = Room(building=building, room_number=room_number, gender=gender, floor=floor_num,
                    capacity=capacity, notes=notes, is_active=True)
        db.session.add(room)
        db.session.flush()

        db.session.add_all([BedAssignment(room_id=room.id, bed_number=bed_num) for bed_num in range(1, capacity + 1)])

        db.session.commit()
        log_operation(current_user, '创建', '宿舍', room.id, f'{room.display_name} {capacity}人间')
        flash(f'宿舍 {room.display_name}（{capacity}人间）已创建', 'success')
        return redirect(url_for('rooms.detail', id=room.id))

    return render_template('dormitory/rooms/form.html', room=None, title='新增宿舍',
                           grades=get_dict_values('grade'), classes=get_dict_values('class'),
                           buildings=get_dict_values('building'), floors=get_dict_values('floor'))


@bp.route('/<int:id>/delete', methods=['POST'])
@perm_required('dormitory.manage')
def delete(id):
    room = Room.query.get_or_404(id)
    occupied = BedAssignment.query.filter(
        BedAssignment.room_id == room.id,
        BedAssignment.student_id.isnot(None)
    ).count()
    if occupied > 0:
        flash(f'{room.room_number} 还有 {occupied} 名学生入住，无法删除', 'danger')
        return redirect(url_for('rooms.list_rooms'))

    BedAssignment.query.filter_by(room_id=room.id).delete()
    db.session.delete(room)
    db.session.commit()
    log_operation(current_user, '删除', '宿舍', room.id, f'{room.display_name}')
    flash(f'宿舍 {room.room_number} 已删除', 'success')
    return redirect(url_for('rooms.list_rooms'))

@bp.route('/assign-visual')
@perm_required('dormitory.manage')
def assign_visual():
    """可视化宿舍分配页面"""
    from app.utils.helpers import get_graduated_grades
    all_grades = get_dict_values('grade')
    graduated = get_graduated_grades()
    grades = [g for g in all_grades if g not in graduated]
    buildings = get_dict_values('building')
    return render_template('dormitory/rooms/assign_visual.html', grades=grades, buildings=buildings)


@bp.route('/assign-data')
@login_required
def assign_data():
    """获取宿舍分配页面所需的所有数据"""
    from flask import jsonify
    from app.utils.helpers import get_graduated_grades
    
    try:
        rooms = Room.query.filter_by(is_active=True).order_by(
            Room.building, Room.floor, Room.room_number
        ).all()
        
        from app.models import Student, StudentAccommodation
        all_grades = get_dict_values('grade')
        graduated = get_graduated_grades()
        grades = [g for g in all_grades if g not in graduated]
        classes_list = get_dict_values('class')
        
        boarding_student_ids = set()
        for acc in StudentAccommodation.query.filter_by(boarding_type='住校').all():
            boarding_student_ids.add(acc.student_id)
        
        classes_data = []
        for grade in grades:
            for cls_name in classes_list:
                male_count = Student.query.filter(
                    Student.grade == grade, 
                    Student.class_name == cls_name,
                    Student.gender == '男',
                    Student.id.in_(boarding_student_ids),
                    ~Student.grade.in_(graduated)
                ).count()
                female_count = Student.query.filter(
                    Student.grade == grade, 
                    Student.class_name == cls_name,
                    Student.gender == '女',
                    Student.id.in_(boarding_student_ids),
                    ~Student.grade.in_(graduated)
                ).count()
                
                if male_count > 0 or female_count > 0:
                    classes_data.append({
                        'grade': grade,
                        'class_name': cls_name,
                        'boarding_male': male_count,
                        'boarding_female': female_count
                    })
        
        rooms_data = []
        for room in rooms:
            rooms_data.append({
                'id': room.id,
                'building': room.building,
                'room_number': room.room_number,
                'floor': room.floor,
                'gender': room.gender,
                'capacity': room.capacity,
                'grade': room.grade,
                'class_name': room.combined_name or room.class_name,
                'combined_class': room.combined_class,
                'is_combined': room.is_combined
            })
        
        return jsonify({
            'rooms': rooms_data,
            'classes': classes_data,
            'buildings': get_dict_values('building'),
            'grades': grades
        })
    
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f'assign-data error: {e}', exc_info=True)
        return jsonify({'error': str(e)}), 500


@bp.route('/assign-room', methods=['POST'])
@perm_required('dormitory.manage')
def assign_room():
    """分配宿舍给班级"""
    from flask import jsonify
    data = request.get_json()
    
    room_id = data.get('room_id')
    grade = data.get('grade')
    class_name = data.get('class_name')
    
    if not room_id or not grade or not class_name:
        return jsonify({'success': False, 'message': '参数不完整'}), 400
    
    room = Room.query.get(room_id)
    if not room:
        return jsonify({'success': False, 'message': '宿舍不存在'}), 404
    
    # 更新宿舍分配信息
    room.grade = grade
    room.class_name = class_name
    # 自动识别合班：class_name 含多个班（+）即视为合班宿舍
    room.combined_details = None
    if class_name and '+' in class_name:
        room.combined_class = class_name
    else:
        room.combined_class = None
    db.session.commit()
    
    return jsonify({'success': True, 'message': '分配成功'})


@bp.route('/unassign-room', methods=['POST'])
@perm_required('dormitory.manage')
def unassign_room():
    """取消宿舍分配"""
    from flask import jsonify
    data = request.get_json()
    
    room_id = data.get('room_id')
    
    if not room_id:
        return jsonify({'success': False, 'message': '参数不完整'}), 400
    
    room = Room.query.get(room_id)
    if not room:
        return jsonify({'success': False, 'message': '宿舍不存在'}), 404
    
    # 取消分配
    room.grade = None
    room.class_name = None
    room.combined_class = None
    room.combined_details = None
    db.session.commit()
    
    return jsonify({'success': True, 'message': '取消成功'})


@bp.route('/set-combined', methods=['POST'])
@perm_required('dormitory.manage')
def set_combined():
    """设置合班宿舍"""
    from flask import jsonify
    data = request.get_json()
    
    room_id = data.get('room_id')
    is_combined = data.get('is_combined', False)
    
    if not room_id:
        return jsonify({'success': False, 'message': '参数不完整'}), 400
    
    room = Room.query.get(room_id)
    if not room:
        return jsonify({'success': False, 'message': '宿舍不存在'}), 404
    
    # 设置合班标记（兼容保留：合班识别现已由 class_name 含多个班自动判定）
    if is_combined:
        room.combined_class = room.class_name if room.class_name and '+' in room.class_name else '合班'
    else:
        room.combined_class = None
        room.combined_details = None
    db.session.commit()
    
    return jsonify({'success': True, 'message': '设置成功'})


@bp.route('/save-assignments', methods=['POST'])
@perm_required('dormitory.manage')
def save_assignments():
    """批量保存所有分配"""
    from flask import jsonify
    data = request.get_json()
    
    assignments = data.get('assignments', [])
    
    # 如果没有分配数据，清空所有宿舍的分配状态
    if not assignments:
        # 清空所有房间分配（包括合班标记）
        rooms = Room.query.filter(Room.grade.isnot(None) | Room.class_name.isnot(None) | Room.combined_class.isnot(None)).all()
        room_count = 0
        for room in rooms:
            room.grade = None
            room.class_name = None
            room.combined_class = None
            room_count += 1
        # 同时清空所有床位
        bed_count = BedAssignment.query.filter(
            BedAssignment.student_id.isnot(None)
        ).update({'student_id': None, 'assigned_by': None, 'assigned_at': None}, synchronize_session=False)
        # 同时清空合班详情残留
        for room in Room.query.filter(Room.combined_details.isnot(None)).all():
            room.combined_details = None
        db.session.commit()
        return jsonify({'success': True, 'message': f'已清空 {room_count} 间房间分配、{bed_count} 个床位'})
    
    # 保存新的分配数据（按房间聚合：同一房间多个班级自动组成合班名，不分主次）
    room_assign = {}
    for assign in assignments:
        rid = assign.get('room_id')
        if not rid:
            continue
        cn = (assign.get('class_name') or '').strip()
        if not cn:
            continue
        grade = (assign.get('grade') or '').strip()
        info = room_assign.setdefault(rid, {'grade': grade, 'classes': []})
        if grade:
            info['grade'] = grade
        for part in cn.split('+'):
            part = part.strip()
            if part and part not in info['classes']:
                info['classes'].append(part)

    count = 0
    for rid, info in room_assign.items():
        room = Room.query.get(rid)
        if not room:
            continue
        # 合班名按班号升序，保证显示规范（不分主次）
        info['classes'].sort(key=_class_sort_key)
        combined_name = '+'.join(info['classes'])
        room.grade = info['grade'] or None
        room.class_name = combined_name or None
        # 手动保存不维护各班级人数份额（自动分配才会写入）
        room.combined_details = None
        # 自动识别合班：超过一个班即自动标记合班
        if '+' in combined_name:
            room.combined_class = combined_name
        else:
            room.combined_class = None
        count += 1

    db.session.commit()
    
    return jsonify({'success': True, 'message': f'已保存 {count} 个分配'})


@bp.route('/class-bed-requirement')
@login_required
def class_bed_requirement():
    """获取某个班级的床位需求和已分配情况"""
    from flask import jsonify
    grade = request.args.get('grade', '')
    class_name = request.args.get('class_name', '')
    
    if not grade or not class_name:
        return jsonify({
            'male': 0, 'female': 0, 'total': 0,
            'assigned_male_rooms': 0, 'assigned_male_beds': 0,
            'assigned_female_rooms': 0, 'assigned_female_beds': 0
        })
    
    # 统计该班级住校学生的性别分布
    boarding_ids = [sa.student_id for sa in StudentAccommodation.query.filter(
        StudentAccommodation.boarding_type == '住校'
    ).all()]
    
    male_count = Student.query.filter(
        Student.grade == grade,
        Student.class_name == class_name,
        Student.gender == '男',
        Student.id.in_(boarding_ids) if boarding_ids else False
    ).count()
    
    female_count = Student.query.filter(
        Student.grade == grade,
        Student.class_name == class_name,
        Student.gender == '女',
        Student.id.in_(boarding_ids) if boarding_ids else False
    ).count()
    
    # 统计该班级已分配的宿舍（含合班宿舍，多个班不分主次）
    assigned_rooms = Room.query.filter(
        Room.is_active == True,
        Room.grade == grade,
        db.or_(
            Room.class_name == class_name,
            Room.class_name.like(class_name + '+%'),
            Room.class_name.like('%+' + class_name + '+%'),
            Room.class_name.like('%+' + class_name),
            Room.combined_class.like(class_name + '+%'),
            Room.combined_class.like('%+' + class_name + '+%'),
            Room.combined_class.like('%+' + class_name)
        )
    ).all()
    
    assigned_male_rooms = 0
    assigned_male_beds = 0
    assigned_female_rooms = 0
    assigned_female_beds = 0
    
    for room in assigned_rooms:
        occupancy = BedAssignment.query.filter(
            BedAssignment.room_id == room.id,
            BedAssignment.student_id.isnot(None)
        ).count()
        
        if room.gender == '男':
            assigned_male_rooms += 1
            assigned_male_beds += occupancy
        else:
            assigned_female_rooms += 1
            assigned_female_beds += occupancy
    
    return jsonify({
        'male': male_count,
        'female': female_count,
        'total': male_count + female_count,
        'assigned_male_rooms': assigned_male_rooms,
        'assigned_male_beds': assigned_male_beds,
        'assigned_female_rooms': assigned_female_rooms,
        'assigned_female_beds': assigned_female_beds
    })


@bp.route('/update-room-assignment', methods=['POST'])
@perm_required('dormitory.manage')
def update_room_assignment():
    """更新单个房间的年级班级分配"""
    from flask import jsonify
    data = request.get_json()
    
    room_id = data.get('room_id')
    grade = data.get('grade', '')
    class_name = data.get('class_name', '')
    
    if not room_id:
        return jsonify({'success': False, 'message': '房间 ID 不能为空'}), 400
    
    room = Room.query.get(room_id)
    if not room:
        return jsonify({'success': False, 'message': '房间不存在'}), 404
    
    # 更新房间分配信息
    room.grade = grade or None
    room.class_name = class_name or None
    # 自动识别合班：class_name 含多个班（+）即视为合班宿舍
    room.combined_details = None
    if class_name and '+' in class_name:
        room.combined_class = class_name
    else:
        room.combined_class = None
    db.session.commit()
    
    return jsonify({'success': True, 'message': '更新成功'})


@bp.route('/batch-setting', methods=['GET', 'POST'])
@perm_required('dormitory.manage')
def batch_setting():
    if request.method == 'POST':
        room_ids = request.form.getlist('room_ids')
        new_gender = request.form.get('gender', '')
        new_capacity = request.form.get('capacity', '')

        if not room_ids:
            flash('请选择宿舍', 'warning')
            return redirect(url_for('rooms.list_rooms'))

        if not new_gender and not new_capacity:
            flash('请至少选择一项要修改的内容（性别或床位数）', 'warning')
            return redirect(url_for('rooms.list_rooms'))

        blocked_rooms = []
        count = 0
        for rid in room_ids:
            room = Room.query.get(int(rid))
            if not room:
                continue
            has_assignment = room.class_name and room.class_name.strip()
            has_beds = BedAssignment.query.filter(
                BedAssignment.room_id == room.id,
                BedAssignment.student_id.isnot(None)
            ).count() > 0
            if has_assignment or has_beds:
                blocked_rooms.append(room.display_name)
                continue
            if new_gender:
                room.gender = new_gender
            if new_capacity:
                cap = int(new_capacity)
                if cap != room.capacity:
                    _adjust_beds(room, room.capacity, cap)
                    room.capacity = cap
            count += 1

        if blocked_rooms:
            db.session.rollback()
            msg = f'{len(blocked_rooms)} 间宿舍无法修改（已分配班级或已安排床位）：{", ".join(blocked_rooms[:10])}'
            if len(blocked_rooms) > 10:
                msg += f' 等{len(blocked_rooms)}间'
            flash(msg, 'danger')
            return redirect(url_for('rooms.list_rooms'))

        db.session.commit()

        msg_parts = []
        if new_gender:
            msg_parts.append(f'性别={new_gender}')
        if new_capacity:
            msg_parts.append(f'床位数={new_capacity}')
        flash(f'已批量设置 {count} 间宿舍（{", ".join(msg_parts)}）', 'success')
        return redirect(url_for('rooms.list_rooms'))

    return redirect(url_for('rooms.list_rooms'))


@bp.route('/batch-add-rooms', methods=['GET', 'POST'])
@perm_required('dormitory.manage')
def batch_add_rooms():
    if request.method == 'POST':
        building = request.form.get('building', '').strip()
        floor = request.form.get('floor', '').strip()
        gender = request.form.get('gender', '男')
        room_count = int(request.form.get('room_count', 0))
        start_room_number = request.form.get('start_room_number', '').strip()
        capacity = int(request.form.get('capacity', 8))

        if not building or not floor or not start_room_number:
            flash('请填写完整信息', 'danger')
            return redirect(url_for('rooms.list_rooms'))

        if room_count <= 0 or room_count > 100:
            flash('房间数量必须在 1-100 之间', 'danger')
            return redirect(url_for('rooms.list_rooms'))

        try:
            start_num = int(start_room_number)
        except ValueError:
            flash('起始房间号必须是数字', 'danger')
            return redirect(url_for('rooms.list_rooms'))

        created_count = 0
        skipped_count = 0

        for i in range(room_count):
            room_num = str(start_num + i)
            existing = Room.query.filter_by(building=building, room_number=room_num).first()
            if existing:
                skipped_count += 1
                continue

            room = Room(
                building=building,
                room_number=room_num,
                gender=gender,
                floor=int(floor),
                capacity=capacity,
                is_active=True
            )
            db.session.add(room)
            db.session.flush()

            for bed_num in range(1, capacity + 1):
                db.session.add(BedAssignment(room_id=room.id, bed_number=bed_num))

            created_count += 1

        db.session.commit()

        msg = f'成功添加 {created_count} 间宿舍'
        if skipped_count > 0:
            msg += f'，跳过 {skipped_count} 间已存在的宿舍'
        flash(msg, 'success')
        return redirect(url_for('rooms.list_rooms'))

    buildings = get_dict_values('building')
    floors = get_dict_values('floor')
    return render_template('dormitory/rooms/batch_add.html', buildings=buildings, floors=floors)


def _adjust_beds(room, old_capacity, new_capacity):
    if new_capacity > old_capacity:
        for bed_num in range(old_capacity + 1, new_capacity + 1):
            existing = BedAssignment.query.filter_by(room_id=room.id, bed_number=bed_num).first()
            if not existing:
                db.session.add(BedAssignment(room_id=room.id, bed_number=bed_num))
    elif new_capacity < old_capacity:
        for bed_num in range(new_capacity + 1, old_capacity + 1):
            bed = BedAssignment.query.filter_by(room_id=room.id, bed_number=bed_num).first()
            if bed:
                if bed.student_id:
                    flash(f'{room.display_name} 的 {bed_num}床 有学生入住，无法删除', 'warning')
                else:
                    db.session.delete(bed)


# ==================== 自动分配宿舍路由 ====================

@bp.route('/assign-auto')
@login_required
@perm_required('dormitory.manage')
def assign_auto():
    """自动分配宿舍向导页面"""
    from app.utils.helpers import get_graduated_grades
    all_grades = get_dict_values('grade')
    graduated = get_graduated_grades()
    grades = [g for g in all_grades if g not in graduated]
    # 按年份降序（新年级在最前/最左）
    grades.sort(key=lambda g: int(''.join(filter(str.isdigit, g)) or '0'), reverse=True)
    
    # 获取各年级各班级的住校生统计
    grade_class_stats = {}
    for grade in grades:
        from sqlalchemy import case
        
        boarding_ids = [sa.student_id for sa in StudentAccommodation.query.filter(
            StudentAccommodation.boarding_type == '住校'
        ).all()]
        
        classes = db.session.query(
            Student.class_name,
            func.count(Student.id).label('count'),
            func.sum(case((Student.gender == '男', 1), else_=0)).label('male'),
            func.sum(case((Student.gender == '女', 1), else_=0)).label('female')
        ).filter(
            Student.grade == grade,
            Student.id.in_(boarding_ids),
            # 排除"已转出"班级和非在读学生
            Student.class_name != '已转出',
            db.or_(
                Student.enrollment_status.is_(None),
                ~Student.enrollment_status.in_(['学籍已转出', '借读后离校'])
            )
        ).filter(~Student.grade.in_(graduated) if graduated else True
        ).group_by(Student.class_name).order_by(Student.class_name).all()
        
        grade_class_stats[grade] = [
            {
                'class_name': c.class_name,
                'count': c.count,
                'male': c.male or 0,
                'female': c.female or 0
            }
            for c in classes
        ]
    
    # 获取楼栋和楼层选项
    buildings = get_dict_values('building')
    floors = get_dict_values('floor')
    
    # 统计住校生总人数（男女分别）供前端第1步显示
    total_boarding_male = 0
    total_boarding_female = 0
    for grade in grades:
        for item in grade_class_stats.get(grade, []):
            total_boarding_male += item.get('male', 0)
            total_boarding_female += item.get('female', 0)

    return render_template('dormitory/rooms/assign_auto.html',
                         grades=grades,
                         grade_class_stats=grade_class_stats,
                         buildings=buildings,
                         floors=floors,
                         total_boarding_male=total_boarding_male,
                         total_boarding_female=total_boarding_female)


@bp.route('/assign-auto/stats')
@login_required
def assign_auto_stats():
    """
    获取班级选择统计数据（后端计算，防止篡改）
    接收: ?keys=grade:class_name:gender,...
    返回: 每个组合的4维度信息 + 汇总
    """
    keys_param = request.args.get('keys', '')
    if not keys_param:
        return jsonify({'success': False, 'error': '参数不完整'})
    
    # 解析选中的班级-性别组合
    selected_keys = []
    for key_str in keys_param.split(','):
        parts = key_str.strip().split(':')
        if len(parts) == 3:
            selected_keys.append({
                'grade': parts[0],
                'class_name': parts[1],
                'gender': parts[2]
            })
    
    if not selected_keys:
        return jsonify({'success': False, 'error': '参数不完整'})
    
    # 从字典表获取有效值
    valid_grades = get_dict_values('grade')
    valid_classes = get_dict_values('class')
    
    # 统计每个组合的真实住校生人数
    details = []
    male_class_count = 0
    male_total = 0
    female_class_count = 0
    female_total = 0
    seen_male_classes = set()
    seen_female_classes = set()
    
    for sk in selected_keys:
        grade = sk['grade']
        class_name = sk['class_name']
        gender = sk['gender']

        # 验证字典表
        if grade not in valid_grades or class_name not in valid_classes:
            continue
        if gender not in ('男', '女'):
            continue
        # 跳过"已转出"班级
        if class_name == '已转出':
            continue

        # 从数据库查询真实人数
        boarding_ids = [sa.student_id for sa in StudentAccommodation.query.filter(
            StudentAccommodation.boarding_type == '住校'
        ).all()]
        student_count = Student.query.filter(
            Student.grade == grade,
            Student.class_name == class_name,
            Student.gender == gender,
            Student.id.in_(boarding_ids) if boarding_ids else False,
            # 排除非在读学生
            db.or_(
                Student.enrollment_status.is_(None),
                ~Student.enrollment_status.in_(['学籍已转出', '借读后离校'])
            )
        ).count()
        
        details.append({
            'grade': grade,
            'class_name': class_name,
            'gender': gender,
            'count': student_count
        })
        
        # 汇总统计
        class_ident = f"{grade}:{class_name}"
        if gender == '男':
            male_total += student_count
            if class_ident not in seen_male_classes:
                seen_male_classes.add(class_ident)
                male_class_count += 1
        else:
            female_total += student_count
            if class_ident not in seen_female_classes:
                seen_female_classes.add(class_ident)
                female_class_count += 1
    
    return jsonify({
        'success': True,
        'details': details,
        'summary': {
            'male_class_count': male_class_count,
            'male_total': male_total,
            'female_class_count': female_class_count,
            'female_total': female_total,
            'total_classes': len(seen_male_classes | seen_female_classes),
            'total_students': male_total + female_total
        }
    })


@bp.route('/assign-auto/preview', methods=['POST'])
@login_required
@perm_required('dormitory.manage')
def assign_auto_preview():
    """预览自动分配方案（不执行）- V7"""
    data = request.json or {}
    
    selected_keys = data.get('selected_keys', [])  # [{grade, class_name, gender}]
    selected_room_ids = data.get('selected_room_ids', [])  # [room_id, ...]
    mode = data.get('mode', 'keep_existing')
    combine_confirmations = data.get('combine_confirmations', [])
    force_full_8 = data.get('force_full_8', False)
    
    if not selected_keys or not selected_room_ids:
        return jsonify({'success': False, 'error': '参数不完整：请选择班级和房间'})
    
    # 检测已有分配的房间
    from app.models import Room as RoomModel, BedAssignment
    rooms_with_assignments = []
    rooms_with_beds = []
    for rid in selected_room_ids:
        room = RoomModel.query.get(rid)
        if not room:
            continue
        if room.class_name and room.class_name.strip():
            rooms_with_assignments.append(f"{room.building} {room.room_number}({room.grade} {room.class_name})")
        bed_count = BedAssignment.query.filter(
            BedAssignment.room_id == rid,
            BedAssignment.student_id.isnot(None)
        ).count()
        if bed_count > 0:
            rooms_with_beds.append(f"{room.building} {room.room_number}({bed_count}床已分配)")
    
    from app.modules.dormitory.services.room_assignment_v6 import auto_assign_preview as do_preview
    
    result = do_preview(
        selected_keys=selected_keys,
        selected_room_ids=selected_room_ids,
        mode=mode,
        occ_ranges=None,
        dry_run=True,
        combine_confirmations=combine_confirmations,
        force_full_8=force_full_8
    )
    
    # 预览模式回滚
    if result['success']:
        db.session.rollback()
    
    # 附加已有分配信息
    result['has_existing'] = len(rooms_with_assignments) > 0 or len(rooms_with_beds) > 0
    result['rooms_with_assignments'] = rooms_with_assignments
    result['rooms_with_beds'] = rooms_with_beds
    
    return jsonify(result)


@bp.route('/assign-auto/execute', methods=['POST'])
@login_required
@perm_required('dormitory.manage')
def assign_auto_execute():
    """执行自动分配 - V7"""
    data = request.json or {}

    selected_keys = data.get('selected_keys', [])  # [{grade, class_name, gender}]
    selected_room_ids = data.get('selected_room_ids', [])  # [room_id, ...]
    mode = data.get('mode', 'keep_existing')
    combine_confirmations = data.get('combine_confirmations', [])
    force_full_8 = data.get('force_full_8', False)
    adjusted_assignments = data.get('adjusted_assignments')  # 用户手动调整后的方案

    if not selected_keys or not selected_room_ids:
        return jsonify({'success': False, 'error': '参数不完整：请选择班级和房间'})

    from app.modules.dormitory.services.room_assignment_v6 import auto_assign_preview as do_preview

    result = do_preview(
        selected_keys=selected_keys,
        selected_room_ids=selected_room_ids,
        mode=mode,
        occ_ranges=None,
        dry_run=False,
        combine_confirmations=combine_confirmations,
        force_full_8=force_full_8,
        adjusted_assignments=adjusted_assignments
    )

    return jsonify(result)


@bp.route('/available-rooms-data')
@login_required
def available_rooms_data():
    """获取可用房间数据（用于自动分配页面）"""
    # 获取所有激活的房间
    rooms = Room.query.filter_by(is_active=True).order_by(
        Room.building, Room.floor, Room.room_number
    ).all()
    
    # 组织成楼栋-楼层-房间的结构
    buildings_data = {}
    
    for room in rooms:
        building = room.building
        floor = room.floor
        
        if building not in buildings_data:
            buildings_data[building] = {}
        
        if floor not in buildings_data[building]:
            buildings_data[building][floor] = []
        
        # 可用床位 = 房间总容量（房间分配阶段不看学生床位）
        buildings_data[building][floor].append({
            'id': room.id,
            'room_number': room.room_number,
            'capacity': room.capacity,
            'occupied': 0,
            'available': room.capacity,
            'gender': room.gender,
            'grade': room.grade,
            'class_name': room.combined_name or room.class_name,
            'is_combined': room.is_combined
        })
    
    return jsonify(buildings_data)


@bp.route('/assign-auto/room-stats', methods=['POST'])
@login_required
def assign_auto_room_stats():
    """
    获取房间选择统计数据（后端计算，防止篡改）
    接收: { room_ids: [1,2,3,...], selected_keys: [{grade,class_name,gender},...] }
    返回: 男女宿舍/床位统计 + 算法估算所需房间数
    """
    import math
    data = request.json or {}
    room_ids = data.get('room_ids', [])
    selected_keys = data.get('selected_keys', [])

    # 统计已选房间的床位数
    if not room_ids:
        male_rooms = female_rooms = male_beds = female_beds = combined_rooms = 0
    else:
        rooms = Room.query.filter(Room.id.in_(room_ids), Room.is_active == True).all()
        if len(rooms) != len(room_ids):
            return jsonify({'success': False, 'error': '部分房间ID无效'})
        male_rooms = female_rooms = male_beds = female_beds = combined_rooms = 0
        for room in rooms:
            if room.gender == '男':
                male_rooms += 1
                male_beds += room.capacity
            elif room.gender == '女':
                female_rooms += 1
                female_beds += room.capacity
            else:
                male_rooms += 1
                male_beds += room.capacity
                female_rooms += 1
                female_beds += room.capacity
            if room.is_combined:
                combined_rooms += 1

    # 统计全部可用（未分配班级）的房间/床位，按性别分类
    avail_male_rooms = avail_female_rooms = 0
    avail_male_beds = avail_female_beds = 0
    all_active = Room.query.filter(Room.is_active == True).all()
    for room in all_active:
        # 跳过已分配班级的房间（这些不可选）
        if room.class_name and room.class_name.strip():
            continue
        if room.gender == '男':
            avail_male_rooms += 1
            avail_male_beds += room.capacity
        elif room.gender == '女':
            avail_female_rooms += 1
            avail_female_beds += room.capacity

    # 根据算法估算所需房间数
    # min = 假设全是8人间所需房间数, max = 假设全是6人间所需房间数
    needed_male_max = needed_male_min = 0
    needed_female_max = needed_female_min = 0
    total_male_students = total_female_students = 0
    if selected_keys:
        valid_grades = get_dict_values('grade')
        valid_classes = get_dict_values('class')
        boarding_ids = [sa.student_id for sa in StudentAccommodation.query.filter(
            StudentAccommodation.boarding_type == '住校'
        ).all()] if selected_keys else []
        # 按班级-性别去重统计人数（同班同性只算一次）
        class_gender_count = {}
        for sk in selected_keys:
            grade = sk.get('grade', '')
            class_name = sk.get('class_name', '')
            gender = sk.get('gender', '')
            if grade not in valid_grades or class_name not in valid_classes:
                continue
            if gender not in ('男', '女'):
                continue
            if class_name == '已转出':
                continue
            key = (grade, class_name, gender)
            if key not in class_gender_count:
                cnt = Student.query.filter(
                    Student.grade == grade,
                    Student.class_name == class_name,
                    Student.gender == gender,
                    Student.id.in_(boarding_ids) if boarding_ids else False,
                    db.or_(
                        Student.enrollment_status.is_(None),
                        ~Student.enrollment_status.in_(['学籍已转出', '借读后离校'])
                    )
                ).count()
                class_gender_count[key] = cnt
        for (grade, class_name, gender), cnt in class_gender_count.items():
            if gender == '男':
                total_male_students += cnt
                needed_male_max += math.ceil(cnt / 6) if cnt > 0 else 0
                needed_male_min += math.ceil(cnt / 8) if cnt > 0 else 0
            else:
                total_female_students += cnt
                needed_female_max += math.ceil(cnt / 6) if cnt > 0 else 0
                needed_female_min += math.ceil(cnt / 8) if cnt > 0 else 0

    # 计算已选房间的6人间/8人间数量，用于更精确的估算
    male_6 = male_8 = female_6 = female_8 = 0
    if room_ids:
        for room in rooms:
            if room.gender == '男':
                if room.capacity == 6:
                    male_6 += 1
                elif room.capacity == 8:
                    male_8 += 1
            elif room.gender == '女':
                if room.capacity == 6:
                    female_6 += 1
                elif room.capacity == 8:
                    female_8 += 1

    # ---- v20260805 实时拥挤度评估（毫秒级，只读不写库） ----
    try:
        from app.modules.dormitory.services.room_assignment_v6 import calc_pressure
        pressure = calc_pressure(selected_keys, room_ids)
    except Exception:
        pressure = {}

    # 计算有效容量（实际可分配床位数 = 所有已选房间的总容量）
    male_effective_cap = male_beds
    female_effective_cap = female_beds

    # 基于实际房间类型计算所需房间数（更精确的估算）
    # 若已有足够8人间支持大班，实际所需会更接近min
    needed_male_est = needed_male_min if total_male_students > 0 else 0
    needed_female_est = needed_female_min if total_female_students > 0 else 0

    return jsonify({
        'success': True,
        'male_rooms': male_rooms,
        'male_beds': male_beds,
        'female_rooms': female_rooms,
        'female_beds': female_beds,
        'total_beds': male_beds + female_beds,
        'combined_rooms': combined_rooms,
        'needed_male_min': needed_male_min,
        'needed_male_max': needed_male_max,
        'needed_female_min': needed_female_min,
        'needed_female_max': needed_female_max,
        'needed_male_est': needed_male_est,
        'needed_female_est': needed_female_est,
        'male_effective_cap': male_effective_cap,
        'female_effective_cap': female_effective_cap,
        'total_male_students': total_male_students,
        'total_female_students': total_female_students,
        'avail_male_rooms': avail_male_rooms,
        'avail_male_beds': avail_male_beds,
        'avail_female_rooms': avail_female_rooms,
        'avail_female_beds': avail_female_beds,
        'male_pressure': pressure.get('male', {}),
        'female_pressure': pressure.get('female', {}),
    })


@bp.route('/report')
@login_required
def report():
    """宿舍报表 → 已合并到 /statistics/?tab=rooms"""
    grade = request.args.get('grade', '')
    url = url_for('statistics.index', tab='rooms')
    if grade:
        url += f'?grade={grade}'
    return redirect(url)


@bp.route('/report/export')
@login_required
def report_export():
    """导出宿舍报表为Excel（保持原路由，由 statistics 页面调用）"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from collections import OrderedDict

    grade_filter = request.args.get('grade', '')

    query = Room.query.filter(
        Room.is_active == True,
        Room.class_name.isnot(None),
        Room.class_name != ''
    )
    if grade_filter:
        query = query.filter_by(grade=grade_filter)

    rooms = query.order_by(
        Room.grade, Room.gender, Room.class_name,
        Room.building, Room.room_number
    ).all()

    tree = OrderedDict()
    class_totals = {}
    for room in rooms:
        g = room.grade or ''
        gender = room.gender or ''
        cn = room.class_name or ''
        if g not in tree:
            tree[g] = OrderedDict()
        if gender not in tree[g]:
            tree[g][gender] = OrderedDict()
        if cn not in tree[g][gender]:
            tree[g][gender][cn] = []
        tree[g][gender][cn].append(room)

    boarding_ids = [sa.student_id for sa in StudentAccommodation.query.filter(
        StudentAccommodation.boarding_type == '住校'
    ).all()]
    
    for g in tree:
        for gender in tree[g]:
            for cn in tree[g][gender]:
                cnt = Student.query.filter(
                    Student.grade == g,
                    Student.class_name == cn,
                    Student.gender == gender,
                    Student.id.in_(boarding_ids) if boarding_ids else False
                ).count()
                class_totals[(g, cn, gender)] = cnt

    wb = Workbook()
    ws = wb.active
    ws.title = '宿舍分配报表'

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11, color='1565C0')
    sub_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    sum_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value='宿舍分配报表')
    c.font = title_font
    c.alignment = Alignment(horizontal='center')
    row += 1

    info = f'已分配 {len(rooms)} 间宿舍 / {sum(r.capacity for r in rooms)} 张床位 / 住校生 {sum(class_totals.values())} 人'
    if grade_filter:
        info += f' / 年级：{grade_filter}'
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value=info)
    c.font = Font(size=10, color='666666')
    c.alignment = Alignment(horizontal='center')
    row += 2

    cols = ['班级', '性别', '住校生', '宿舍楼', '房间号', '床位数', '合班标记']
    col_widths = [10, 6, 9, 18, 9, 9, 14]

    for grade, genders in tree.items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=f'▌ {grade}')
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal='left')
        row += 1

        for gender, classes in genders.items():
            gender_label = '男生' if gender == '男' else '女生'
            gender_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid') if gender == '男' else PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            c = ws.cell(row=row, column=1, value=f'{gender_label}')
            c.font = subtitle_font
            c.fill = gender_fill
            row += 1

            for ci, (col_name, width) in enumerate(zip(cols, col_widths), 1):
                c = ws.cell(row=row, column=ci, value=col_name)
                c.font = header_font
                c.fill = header_fill
                c.border = thin_border
                c.alignment = center_align
                ws.column_dimensions[get_column_letter(ci)].width = width
            row += 1

            for class_name, room_list in classes.items():
                boarders = class_totals.get((grade, class_name, gender), 0)
                room_count = len(room_list)
                start_row_class = row

                for ri, room in enumerate(room_list):
                    r = row
                    if ri == 0:
                        c = ws.cell(row=r, column=1, value=grade + class_name)
                        ws.merge_cells(start_row=r, start_column=1, end_row=r + room_count - 1, end_column=1)
                        c.font = Font(bold=True)
                        c.alignment = center_align
                        c = ws.cell(row=r, column=2, value=gender)
                        ws.merge_cells(start_row=r, start_column=2, end_row=r + room_count - 1, end_column=2)
                        c.font = Font(bold=True)
                        c.alignment = center_align
                        c = ws.cell(row=r, column=3, value=boarders)
                        ws.merge_cells(start_row=r, start_column=3, end_row=r + room_count - 1, end_column=3)
                        c.font = Font(bold=True)
                        c.alignment = center_align

                    ws.cell(row=r, column=4, value=room.building).alignment = left_align
                    ws.cell(row=r, column=5, value=room.room_number).alignment = center_align
                    ws.cell(row=r, column=6, value=room.capacity).alignment = center_align
                    combined = room.combined_class if room.combined_class and room.combined_class.strip() else ''
                    ws.cell(row=r, column=7, value=combined).alignment = center_align

                    for ci in range(1, 8):
                        ws.cell(row=r, column=ci).border = thin_border
                    row += 1

                # 小计行
                cls_beds = sum(r.capacity for r in room_list)
                c = ws.cell(row=row, column=5, value=f'小计：{room_count}间')
                c.font = sum_font
                c.alignment = Alignment(horizontal='right')
                c = ws.cell(row=row, column=6, value=f'{cls_beds}床')
                c.font = sum_font
                c.alignment = center_align
                for ci in range(1, 8):
                    ws.cell(row=row, column=ci).border = thin_border
                    ws.cell(row=row, column=ci).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                row += 1

            row += 1  # 性别间空行

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from datetime import datetime
    from flask import send_file
    from app.models import OperationLog
    from app.extensions import db
    import json

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'宿舍分配报表_{grade_filter or "全部"}_{timestamp}.xlsx'

    try:
        log_detail = {
            'columns': ['班级', '性别', '住校生', '宿舍楼', '房间号', '床位数', '合班标记'],
            'record_count': len(rooms),
            'file_name': filename,
            'filters': {'grade': grade_filter}
        }
        log = OperationLog(
            user_id=current_user.id,
            action='导出',
            target_type='宿舍分配报表',
            module='dormitory',
            detail=json.dumps(log_detail, ensure_ascii=False)
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        db.session.rollback()

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@bp.route('/swap-data')
@login_required
def swap_data():
    """获取宿舍互换所需的房间列表数据"""
    rooms = Room.query.filter_by(is_active=True).order_by(
        Room.gender, Room.building, Room.floor, Room.room_number
    ).all()

    # 批量获取入住人数
    occupancy_data = db.session.query(
        BedAssignment.room_id,
        func.count(BedAssignment.id).label('count')
    ).filter(
        BedAssignment.student_id.isnot(None)
    ).group_by(BedAssignment.room_id).all()
    occupancy_map = {row.room_id: row.count for row in occupancy_data}

    result = []
    for room in rooms:
        occ = occupancy_map.get(room.id, 0)
        cn = room.combined_name or room.class_name
        class_info = ''
        if room.grade and cn:
            class_info = f"{room.grade} {cn}"
        elif room.grade:
            class_info = room.grade
        elif cn:
            class_info = cn

        result.append({
            'id': room.id,
            'building': room.building,
            'room_number': room.room_number,
            'floor': room.floor,
            'gender': room.gender,
            'capacity': room.capacity,
            'occupancy': occ,
            'grade': room.grade or '',
            'class_name': cn or '',
            'combined_class': room.combined_class or '',
            'is_combined': room.is_combined,
            'class_info': class_info,
        })

    return jsonify({'success': True, 'rooms': result})


@bp.route('/swap', methods=['POST'])
@perm_required('dormitory.manage')
def swap_rooms():
    """交换两个宿舍的班级分配和学生床位"""
    data = request.get_json()
    room_a_id = data.get('room_a_id')
    room_b_id = data.get('room_b_id')

    if not room_a_id or not room_b_id:
        return jsonify({'success': False, 'message': '请选择两个宿舍'}), 400
    if room_a_id == room_b_id:
        return jsonify({'success': False, 'message': '不能选择同一个宿舍'}), 400

    room_a = Room.query.get(room_a_id)
    room_b = Room.query.get(room_b_id)
    if not room_a or not room_b:
        return jsonify({'success': False, 'message': '宿舍不存在'}), 404
    if not room_a.is_active or not room_b.is_active:
        return jsonify({'success': False, 'message': '宿舍已停用'}), 400

    # 性别校验：必须同性别
    if room_a.gender != room_b.gender:
        return jsonify({'success': False, 'message': f'性别不同，无法互换（{room_a.gender}生 vs {room_b.gender}生）'}), 400

    # 获取两间房的入住学生（按床位号顺序）
    beds_a = BedAssignment.query.filter_by(room_id=room_a.id).order_by(BedAssignment.bed_number).all()
    beds_b = BedAssignment.query.filter_by(room_id=room_b.id).order_by(BedAssignment.bed_number).all()
    students_a = [b.student_id for b in beds_a if b.student_id]
    students_b = [b.student_id for b in beds_b if b.student_id]

    # 容量校验：A的入住人数不能超过B的床位数，反之亦然
    if len(students_a) > room_b.capacity:
        return jsonify({'success': False, 'message': f'{room_a.building}{room_a.room_number}有{len(students_a)}人，超过{room_b.building}{room_b.room_number}的{room_b.capacity}个床位'}), 400
    if len(students_b) > room_a.capacity:
        return jsonify({'success': False, 'message': f'{room_b.building}{room_b.room_number}有{len(students_b)}人，超过{room_a.building}{room_a.room_number}的{room_a.capacity}个床位'}), 400

    # 记录互换前的信息用于日志
    a_info = f"{room_a.building}{room_a.room_number}({room_a.grade or ''} {room_a.class_name or '未分配'}, {len(students_a)}人)"
    b_info = f"{room_b.building}{room_b.room_number}({room_b.grade or ''} {room_b.class_name or '未分配'}, {len(students_b)}人)"

    try:
        # 1. 交换房间属性（年级、班级、合班标记）
        room_a.grade, room_b.grade = room_b.grade, room_a.grade
        room_a.class_name, room_b.class_name = room_b.class_name, room_a.class_name
        room_a.combined_class, room_b.combined_class = room_b.combined_class, room_a.combined_class

        # 2. 交换学生床位
        # 清空两个房间的所有学生
        for bed in beds_a:
            bed.student_id = None
            bed.assigned_by = None
        for bed in beds_b:
            bed.student_id = None
            bed.assigned_by = None
        db.session.flush()

        # A的学生 → B的床位（按床位号顺序依次填入）
        for i, sid in enumerate(students_a):
            if i < len(beds_b):
                beds_b[i].student_id = sid
                beds_b[i].assigned_by = current_user.id

        # B的学生 → A的床位（按床位号顺序依次填入）
        for i, sid in enumerate(students_b):
            if i < len(beds_a):
                beds_a[i].student_id = sid
                beds_a[i].assigned_by = current_user.id

        db.session.commit()

        log_operation(current_user, '互换', '宿舍', None,
                      f'{a_info} ⇄ {b_info}')

        return jsonify({
            'success': True,
            'message': f'互换成功：{a_info} ⇄ {b_info}'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'互换失败：{str(e)}'}), 500


