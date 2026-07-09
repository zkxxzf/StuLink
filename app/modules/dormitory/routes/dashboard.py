# StuLink v1.6.1 2026-07-09
# Copyright (c) 2026 zkxxzf. CC BY-NC 4.0
import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from app.models import Student, Room, BedAssignment, User, StudentAccommodation
from app.extensions import db
from sqlalchemy import func, case, text
from app.utils.helpers import get_graduated_grades, get_dict_values
from app.utils.decorators import perm_required
from app.utils.helpers import log_operation

bp = Blueprint('dashboard', __name__, url_prefix='/dashboard')


@bp.route('/')
@login_required
def index():
    graduated = get_graduated_grades()
    base_q = db.session.query(
        func.count(Student.id).label('total'),
        func.sum(case((Student.gender == '男', 1), else_=0)).label('male'),
        func.sum(case((Student.gender == '女', 1), else_=0)).label('female')
    )
    if graduated:
        base_q = base_q.filter(~Student.grade.in_(graduated))
    student_stats = base_q.one()

    # 查询 dormitory.db 中的住宿统计
    boarding_count = StudentAccommodation.query.filter(
        StudentAccommodation.boarding_type == '住校'
    ).count()
    day_count = StudentAccommodation.query.filter(
        StudentAccommodation.boarding_type == '走读'
    ).count()

    total_rooms = db.session.query(func.count(Room.id)).filter(Room.is_active == True).scalar()
    
    active_room_ids = db.session.query(Room.id).filter(Room.is_active == True).all()
    active_room_ids = [r[0] for r in active_room_ids]
    
    if active_room_ids:
        total_beds = db.session.query(func.count(BedAssignment.id)).filter(
            BedAssignment.room_id.in_(active_room_ids)
        ).scalar()
        assigned_beds = db.session.query(func.count(BedAssignment.id)).filter(
            BedAssignment.room_id.in_(active_room_ids),
            BedAssignment.student_id.isnot(None)
        ).scalar()
    else:
        total_beds = 0
        assigned_beds = 0
    
    total_users = db.session.query(func.count(User.id)).filter(User.is_active == True).scalar()

    stats = {
        'total_students': student_stats.total or 0,
        'male_students': student_stats.male or 0,
        'female_students': student_stats.female or 0,
        'boarding_students': boarding_count,
        'day_students': day_count,
        'total_rooms': total_rooms or 0,
        'assigned_beds': assigned_beds or 0,
        'total_beds': total_beds or 0,
        'total_users': total_users or 0,
    }

    grade_query = db.session.query(
        Student.grade,
        func.count(Student.id).label('total'),
        func.sum(case((Student.gender == '男', 1), else_=0)).label('male'),
        func.sum(case((Student.gender == '女', 1), else_=0)).label('female')
    ).group_by(Student.grade).order_by(Student.grade).all()

    grade_accomm_map = {}
    for g in grade_query:
        grade = g.grade
        grade_student_ids = [s.id for s in Student.query.filter_by(grade=grade).all()]
        grade_boarding = StudentAccommodation.query.filter(
            StudentAccommodation.student_id.in_(grade_student_ids),
            StudentAccommodation.boarding_type == '住校'
        ).count()
        grade_day = StudentAccommodation.query.filter(
            StudentAccommodation.student_id.in_(grade_student_ids),
            StudentAccommodation.boarding_type == '走读'
        ).count()
        grade_accomm_map[grade] = {'boarding': grade_boarding, 'day': grade_day}

    grade_stats = []
    for g in grade_query:
        acc_data = grade_accomm_map.get(g.grade, {'boarding': 0, 'day': 0})
        grade_stats.append({
            'grade': g.grade,
            'total': g.total,
            'male': g.male,
            'female': g.female,
            'boarding': acc_data['boarding'],
            'day': acc_data['day'],
        })

    return render_template('dormitory/dashboard/index.html', stats=stats, grade_stats=grade_stats)


@bp.route('/search')
@login_required
def search_student_accommodation():
    query = Student.query
    graduated = get_graduated_grades()
    if graduated:
        query = query.filter(~Student.grade.in_(graduated))

    if current_user.role == 'homeroom_teacher':
        query = query.filter_by(grade=current_user.grade, class_name=current_user.class_name)
    elif current_user.role == 'grade_leader':
        query = query.filter_by(grade=current_user.grade)

    name = request.args.get('name', '').strip()
    student_number = request.args.get('student_number', '').strip()
    grade = request.args.get('grade', '')
    class_name = request.args.get('class_name', '')
    gender = request.args.get('gender', '')
    boarding_type = request.args.get('boarding_type', '')
    day_student_type = request.args.get('day_student_type', '')
    subject_selection = request.args.get('subject_selection', '')
    room_number = request.args.get('room_number', '').strip()

    if name:
        query = query.filter(Student.name.contains(name))
    if student_number:
        query = query.filter(Student.student_number.contains(student_number))
    if grade:
        query = query.filter_by(grade=grade)
    if class_name:
        query = query.filter_by(class_name=class_name)
    if gender:
        query = query.filter_by(gender=gender)
    if boarding_type:
        acc_ids = [sa.student_id for sa in StudentAccommodation.query.filter_by(boarding_type=boarding_type).all()]
        if acc_ids:
            query = query.filter(Student.id.in_(acc_ids))
        else:
            query = query.filter(Student.id == -1)
    if day_student_type:
        acc_ids = [sa.student_id for sa in StudentAccommodation.query.filter_by(day_student_type=day_student_type).all()]
        if acc_ids:
            query = query.filter(Student.id.in_(acc_ids))
        else:
            query = query.filter(Student.id == -1)
    if subject_selection:
        query = query.filter_by(subject_selection=subject_selection)
    if room_number:
        bed_sub = db.session.query(BedAssignment.student_id).join(BedAssignment.room).filter(
            Room.room_number.contains(room_number)
        ).filter(BedAssignment.student_id.isnot(None)).all()
        bed_ids = [b[0] for b in bed_sub if b[0]]
        if bed_ids:
            query = query.filter(Student.id.in_(bed_ids))
        else:
            query = query.filter(Student.id == -1)

    page = request.args.get('page', 1, type=int)
    per_page = 50

    pagination = query.order_by(Student.grade, Student.class_name, Student.student_number).paginate(
        page=page, per_page=per_page, error_out=False
    )
    students = pagination.items

    grades = get_dict_values('grade')
    grades = [g for g in grades if g not in graduated]
    classes = get_dict_values('class')
    boarding_types = get_dict_values('boarding_type')
    day_student_types = get_dict_values('day_student_type')
    subjects = get_dict_values('subject')
    buildings = get_dict_values('building')
    return render_template('dormitory/dashboard/student_accommodation.html',
                           students=students,
                           grades=grades,
                           classes=classes,
                           boarding_types=boarding_types,
                           day_student_types=day_student_types,
                           subjects=subjects,
                           buildings=buildings,
                           pagination=pagination)


@bp.route('/search/export', methods=['GET', 'POST'])
@bp.route('/export', methods=['GET', 'POST'])
@login_required
def export_student_accommodation():
    from app.utils.export_helpers import do_export_student_accommodation
    if request.method == 'POST':
        return do_export_student_accommodation(request.form)
    return do_export_student_accommodation(request.args)


@bp.route('/search/batch-edit-dormitory', methods=['POST'])
@perm_required('dormitory.manage')
def batch_edit_dormitory():
    ids = request.form.getlist('student_ids')
    if not ids:
        flash('未选择任何学生', 'warning')
        return redirect(url_for('dashboard.search_student_accommodation'))

    try:
        id_list = [int(i) for i in ids]
    except ValueError:
        flash('参数错误', 'danger')
        return redirect(url_for('dashboard.search_student_accommodation'))

    students = Student.query.filter(Student.id.in_(id_list)).all()
    if not students:
        flash('未找到选中的学生', 'warning')
        return redirect(url_for('dashboard.search_student_accommodation'))

    acc_fields = {
        'boarding_type': '住校/走读',
        'day_student_type': '出门权限',
        'textbook': '课本',
        'teacher_notes': '班主任备注',
    }
    student_fields = {
        'subject_selection': '选科',
    }

    updated_fields = []

    for field, label in acc_fields.items():
        val = request.form.get(field, '').strip()
        if val:
            for s in students:
                acc = StudentAccommodation.query.filter_by(student_id=s.id).first()
                if not acc:
                    acc = StudentAccommodation(student_id=s.id)
                    db.session.add(acc)
                setattr(acc, field, val)
            updated_fields.append(f'{label}={val}')

    for field, label in student_fields.items():
        val = request.form.get(field, '').strip()
        if val:
            for s in students:
                setattr(s, field, val)
            updated_fields.append(f'{label}={val}')

    if updated_fields:
        db.session.commit()
        log_operation(current_user, '批量修改', '学生', None,
                       f'{len(students)}名学生：{", ".join(updated_fields)}')
        flash(f'已更新{len(students)}名学生的：{", ".join(updated_fields)}', 'success')
    else:
        flash('未选择任何要修改的字段', 'warning')

    return redirect(url_for('dashboard.search_student_accommodation'))


@bp.route('/search/import-template')
@login_required
def download_dorm_import_template():
    """下载宿舍信息导入模板"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '宿舍信息导入模板'

    headers = ['学号', '姓名', '住校/走读', '出门权限', '宿舍号', '床位号']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    required_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    # 学号为必填（橙色），其余选填
    required_cols = {0}
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = required_fill if (col_idx - 1) in required_cols else header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    widths = [15, 10, 12, 14, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 示例数据
    ws.append(['20251001', '张三', '住校', '', '东201', '1'])
    ws.append(['20251002', '李四', '走读', '半走读', '', ''])

    # 说明行
    ws.merge_cells('A3:F3')
    instr = ws.cell(row=3, column=1,
                    value='说明：学号必填(橙色) | 住校/走读填"住校"或"走读" | 宿舍号格式如"东201"（楼栋+房间号） | 床位号为数字 | 宿舍号和床位号可单独填写 | 留空的字段不修改 | 删除本行和示例后上传')
    instr.font = Font(color='FF0000', bold=True, size=10)
    instr.alignment = Alignment(horizontal='left')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='宿舍信息导入模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/search/import', methods=['POST'])
@perm_required('dormitory.manage')
def import_dormitory_info():
    """通过Excel批量导入宿舍信息：住校/走读、出门权限、宿舍号、床位号"""
    file = request.files.get('file')
    if not file or not file.filename:
        flash('请选择要上传的Excel文件', 'danger')
        return redirect(url_for('dashboard.search_student_accommodation'))

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash(f'文件格式不正确，请上传 .xlsx 格式的Excel文件（当前文件：{file.filename}）', 'danger')
        return redirect(url_for('dashboard.search_student_accommodation'))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # 读取表头
        header_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                header_map[str(val).strip()] = col

        if not header_map:
            flash('Excel文件表头为空，请确保第一行为表头', 'danger')
            return redirect(url_for('dashboard.search_student_accommodation'))

        field_mapping = {
            '学号': 'student_number', '姓名': 'name',
            '住校/走读': 'boarding_type', '住宿类型': 'boarding_type',
            '出门权限': 'day_student_type',
            '宿舍号': 'room_number', '房间号': 'room_number',
            '床位号': 'bed_number', '床位': 'bed_number',
        }
        col_map = {}
        for excel_name, model_field in field_mapping.items():
            if excel_name in header_map and model_field not in col_map:
                col_map[model_field] = header_map[excel_name]

        if 'student_number' not in col_map:
            flash('Excel缺少必填列：学号，请使用系统下载的导入模板', 'danger')
            return redirect(url_for('dashboard.search_student_accommodation'))

        # 预加载所有学生（按学号索引）
        all_snums = set()
        for row_idx in range(2, ws.max_row + 1):
            snum_val = ws.cell(row=row_idx, column=col_map['student_number']).value
            if snum_val:
                all_snums.add(str(snum_val).strip())

        students_map = {}
        if all_snums:
            for s in Student.query.filter(Student.student_number.in_(all_snums)).all():
                students_map[s.student_number] = s

        # 预加载所有活跃房间（按 building+room_number 索引）
        rooms_map = {}
        for r in Room.query.filter(Room.is_active == True).all():
            key = f"{r.building}{r.room_number}"
            rooms_map[key] = r

        errors = []
        updated_count = 0
        bed_assigned_count = 0

        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for model_field, col_idx in col_map.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data[model_field] = str(val).strip() if val is not None else ''

            snum = row_data.get('student_number', '')
            if not snum:
                continue

            student = students_map.get(snum)
            if not student:
                errors.append(f'第{row_idx}行：学号"{snum}"不存在')
                continue

            row_updated = False
            name_in_excel = row_data.get('name', '')
            if name_in_excel and student.name != name_in_excel:
                errors.append(f'第{row_idx}行：学号"{snum}"姓名不匹配（Excel:{name_in_excel} / 系统:{student.name}）')
                continue

            # 1. 更新住校/走读（写入 StudentAccommodation）
            bt = row_data.get('boarding_type', '')
            if bt:
                if bt not in ('住校', '走读'):
                    errors.append(f'第{row_idx}行：住校/走读"{bt}"无效（应为"住校"或"走读"）')
                else:
                    acc = StudentAccommodation.query.filter_by(student_id=student.id).first()
                    if not acc:
                        acc = StudentAccommodation(student_id=student.id)
                        db.session.add(acc)
                    if acc.boarding_type != bt:
                        acc.boarding_type = bt
                        row_updated = True

            # 2. 更新出门权限（写入 StudentAccommodation）
            dst = row_data.get('day_student_type', '')
            if dst:
                acc = StudentAccommodation.query.filter_by(student_id=student.id).first()
                if not acc:
                    acc = StudentAccommodation(student_id=student.id)
                    db.session.add(acc)
                if acc.day_student_type != dst:
                    acc.day_student_type = dst
                    row_updated = True

            # 3. 分配宿舍号 + 床位号
            room_str = row_data.get('room_number', '')
            bed_str = row_data.get('bed_number', '')
            if room_str or bed_str:
                if room_str and bed_str:
                    try:
                        bed_num = int(bed_str)
                    except ValueError:
                        errors.append(f'第{row_idx}行：床位号"{bed_str}"不是有效数字')
                        continue

                    room = rooms_map.get(room_str)
                    if not room:
                        errors.append(f'第{row_idx}行：宿舍号"{room_str}"不存在')
                        continue

                    if bed_num < 1 or bed_num > room.capacity:
                        errors.append(f'第{row_idx}行：床位号{bed_num}超出房间容量(1-{room.capacity})')
                        continue

                    # 检查性别匹配
                    if student.gender != room.gender:
                        errors.append(f'第{row_idx}行：学生性别({student.gender})与房间性别({room.gender})不匹配')
                        continue

                    # 查找目标床位
                    target_bed = BedAssignment.query.filter_by(
                        room_id=room.id, bed_number=bed_num
                    ).first()

                    if target_bed and target_bed.student_id and target_bed.student_id != student.id:
                        other = Student.query.get(target_bed.student_id)
                        other_name = other.name if other else '未知'
                        errors.append(f'第{row_idx}行：{room_str}的{bed_num}床已被{other_name}占用')
                        continue

                    # 清除学生原有床位
                    old_bed = BedAssignment.query.filter_by(student_id=student.id).first()
                    if old_bed and old_bed.id != (target_bed.id if target_bed else None):
                        old_bed.student_id = None
                        old_bed.assigned_by = None
                        db.session.add(old_bed)

                    # 分配新床位
                    if target_bed:
                        target_bed.student_id = student.id
                        target_bed.assigned_by = current_user.id
                        db.session.add(target_bed)
                    else:
                        new_bed = BedAssignment(
                            room_id=room.id,
                            bed_number=bed_num,
                            student_id=student.id,
                            assigned_by=current_user.id
                        )
                        db.session.add(new_bed)
                    bed_assigned_count += 1
                    row_updated = True
                elif room_str:
                    room = rooms_map.get(room_str)
                    if not room:
                        errors.append(f'第{row_idx}行：宿舍号"{room_str}"不存在')
                        continue
                    old_bed = BedAssignment.query.filter_by(student_id=student.id).first()
                    if old_bed:
                        old_bed.student_id = None
                        old_bed.assigned_by = None
                        db.session.add(old_bed)
                        row_updated = True

            if row_updated:
                updated_count += 1

        db.session.commit()

        summary_parts = [f'更新{updated_count}名学生']
        if bed_assigned_count:
            summary_parts.append(f'分配床位{bed_assigned_count}个')
        if errors:
            summary_parts.append(f'{len(errors)}条错误')
        summary = '，'.join(summary_parts)

        log_operation(current_user, '导入', '宿舍信息', None, summary)

        if errors:
            flash(f'导入完成：{summary}', 'warning')
            flash('错误详情：\n' + '\n'.join(errors[:20]), 'info')
        else:
            flash(f'导入完成：{summary}', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'导入失败：{str(e)}', 'danger')

    return redirect(url_for('dashboard.search_student_accommodation'))


