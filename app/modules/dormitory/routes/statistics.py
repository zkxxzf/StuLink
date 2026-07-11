# StuLink v1.6.1 2026-07-09
# Copyright (c) 2026 zkxxzf. Apache License 2.0
from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file
from markupsafe import Markup
from flask_login import login_required, current_user
from app.models import Student, Room, BedAssignment, UserClassLink, StudentAccommodation
from app.extensions import db
from app.utils.helpers import get_dict_values, log_operation, get_graduated_grades
from app.utils.decorators import perm_required
import io
import uuid
import time

bp = Blueprint('statistics', __name__, url_prefix='/statistics')

SCOPE_CLASS = 'class'    # 班主任：只看所管班级
SCOPE_GRADE = 'grade'    # 年级长：只看所管年级
SCOPE_SCHOOL = 'school'  # 全校组/admin：看全部


def _get_scope():
    """获取当前用户的权限范围"""
    pg = current_user.permission_group
    if not pg:
        return SCOPE_SCHOOL, None
    return pg.scope_type, current_user.grade


def _get_user_class_links():
    """获取当前班主任的所有班级关联"""
    return UserClassLink.query.filter_by(user_id=current_user.id).all()


def _build_per_class_stats(filter_grade=None, filter_classes=None):
    """按年级+班级统计"""
    graduated = get_graduated_grades()
    q = db.session.query(
        Student.grade, Student.class_name
    ).distinct().order_by(Student.grade, Student.class_name)
    if graduated:
        q = q.filter(~Student.grade.in_(graduated))
    if filter_grade:
        q = q.filter(Student.grade == filter_grade)

    results = q.all()
    if filter_classes:
        allowed = {(g, c) for g, c in filter_classes}
        results = [r for r in results if (r[0], r[1]) in allowed]

    stats = []
    for grade, class_name in results:
        graduated = get_graduated_grades()
        qs = Student.query.filter_by(grade=grade, class_name=class_name)
        if graduated:
            qs = qs.filter(~Student.grade.in_(graduated))
        total = qs.count()
        male = qs.filter_by(gender='男').count()
        female = qs.filter_by(gender='女').count()
        
        student_ids = [s.id for s in qs.all()]
        boarding = StudentAccommodation.query.filter(
            StudentAccommodation.student_id.in_(student_ids),
            StudentAccommodation.boarding_type == '住校'
        ).count()
        day_student = StudentAccommodation.query.filter(
            StudentAccommodation.student_id.in_(student_ids),
            StudentAccommodation.boarding_type == '走读'
        ).count()
        
        male_ids = [s.id for s in qs.filter_by(gender='男').all()]
        male_boarding = StudentAccommodation.query.filter(
            StudentAccommodation.student_id.in_(male_ids),
            StudentAccommodation.boarding_type == '住校'
        ).count()
        male_day = StudentAccommodation.query.filter(
            StudentAccommodation.student_id.in_(male_ids),
            StudentAccommodation.boarding_type == '走读'
        ).count()
        
        female_ids = [s.id for s in qs.filter_by(gender='女').all()]
        female_boarding = StudentAccommodation.query.filter(
            StudentAccommodation.student_id.in_(female_ids),
            StudentAccommodation.boarding_type == '住校'
        ).count()
        female_day = StudentAccommodation.query.filter(
            StudentAccommodation.student_id.in_(female_ids),
            StudentAccommodation.boarding_type == '走读'
        ).count()

        stats.append({
            'grade': grade, 'class_name': class_name,
            'total': total, 'male': male, 'female': female,
            'boarding': boarding, 'male_boarding': male_boarding,
            'female_boarding': female_boarding,
            'day_student': day_student, 'male_day': male_day, 'female_day': female_day,
        })
    return stats


def _build_per_grade_stats(filter_grade=None):
    """按年级汇总"""
    graduated = get_graduated_grades()
    q = db.session.query(Student.grade).distinct().order_by(Student.grade)
    if graduated:
        q = q.filter(~Student.grade.in_(graduated))
    if filter_grade:
        q = q.filter(Student.grade == filter_grade)
    grades = [r[0] for r in q.all()]

    result = []
    for grade in grades:
        qs = Student.query.filter_by(grade=grade)
        if graduated:
            qs = qs.filter(~Student.grade.in_(graduated))
        total = qs.count()
        male = qs.filter_by(gender='男').count()
        female = qs.filter_by(gender='女').count()
        
        student_ids = [s.id for s in qs.all()]
        boarding = StudentAccommodation.query.filter(
            StudentAccommodation.student_id.in_(student_ids),
            StudentAccommodation.boarding_type == '住校'
        ).count()
        
        male_ids = [s.id for s in qs.filter_by(gender='男').all()]
        male_boarding = StudentAccommodation.query.filter(
            StudentAccommodation.student_id.in_(male_ids),
            StudentAccommodation.boarding_type == '住校'
        ).count()
        
        female_ids = [s.id for s in qs.filter_by(gender='女').all()]
        female_boarding = StudentAccommodation.query.filter(
            StudentAccommodation.student_id.in_(female_ids),
            StudentAccommodation.boarding_type == '住校'
        ).count()
        
        class_count = db.session.query(Student.class_name).filter_by(grade=grade).distinct().count()

        result.append({
            'grade': grade, 'class_count': class_count,
            'total': total, 'male': male, 'female': female,
            'boarding': boarding, 'male_boarding': male_boarding,
            'female_boarding': female_boarding,
        })
    return result


def _build_school_stats():
    """全校汇总"""
    graduated = get_graduated_grades()
    qs = Student.query
    if graduated:
        qs = qs.filter(~Student.grade.in_(graduated))
    total = qs.count()
    male = qs.filter_by(gender='男').count()
    female = qs.filter_by(gender='女').count()
    
    student_ids = [s.id for s in qs.all()]
    boarding = StudentAccommodation.query.filter(
        StudentAccommodation.student_id.in_(student_ids),
        StudentAccommodation.boarding_type == '住校'
    ).count()
    
    male_ids = [s.id for s in qs.filter_by(gender='男').all()]
    male_boarding = StudentAccommodation.query.filter(
        StudentAccommodation.student_id.in_(male_ids),
        StudentAccommodation.boarding_type == '住校'
    ).count()
    
    female_ids = [s.id for s in qs.filter_by(gender='女').all()]
    female_boarding = StudentAccommodation.query.filter(
        StudentAccommodation.student_id.in_(female_ids),
        StudentAccommodation.boarding_type == '住校'
    ).count()
    
    grade_count = db.session.query(Student.grade).distinct().count()
    class_count = db.session.query(Student.grade, Student.class_name).distinct().count()

    return {
        'grade_count': grade_count, 'class_count': class_count,
        'total': total, 'male': male, 'female': female,
        'boarding': boarding, 'male_boarding': male_boarding,
        'female_boarding': female_boarding,
    }


def _dorm_stats():
    total_rooms = Room.query.filter_by(is_active=True).count()
    occupied_beds = BedAssignment.query.filter(BedAssignment.student_id.isnot(None)).count()
    total_beds = BedAssignment.query.count()
    return {
        'total_rooms': total_rooms,
        'total_beds': total_beds,
        'occupied_beds': occupied_beds,
        'empty_beds': total_beds - occupied_beds,
        'occupancy_rate': round(occupied_beds / total_beds * 100, 1) if total_beds else 0,
    }


@bp.route('/')
@login_required
def index():
    scope_type, user_grade = _get_scope()
    tab = request.args.get('tab', scope_type)  # 默认选用户范围对应的tab
    if tab not in ('school', 'grade', 'class', 'import', 'rooms'):
        tab = 'school'
    sel_grade = request.args.get('grade', user_grade or '')

    # 班主任：只允许看 class tab
    if scope_type == SCOPE_CLASS:
        tab = 'class'
        links = _get_user_class_links()
        allowed = [(l.grade, l.class_name) for l in links]
        per_class_stats = _build_per_class_stats(filter_classes=allowed)
        per_grade_stats = []
        school_stats = {}
        grade_options = list(set(l.grade for l in links))
    # 年级长：看 grade 或 class tab，限制年级
    elif scope_type == SCOPE_GRADE:
        if tab == 'school':
            tab = 'grade'
        if not sel_grade:
            sel_grade = user_grade or ''
        per_class_stats = _build_per_class_stats(filter_grade=sel_grade)
        per_grade_stats = _build_per_grade_stats(filter_grade=user_grade)
        school_stats = _build_school_stats() if tab == 'school' else {}
        grade_options = [user_grade] if user_grade else []
    else:
        # 全校组/admin：全部数据
        per_class_stats = _build_per_class_stats(filter_grade=sel_grade if tab == 'class' and sel_grade else None)
        per_grade_stats = _build_per_grade_stats()
        school_stats = _build_school_stats()
        grade_options = sorted(get_dict_values('grade'), reverse=True)

    # 宿舍分配明细（原 /rooms/report）
    room_tree = {}
    room_class_totals = {}
    room_total_beds = 0
    room_total_boarders = 0
    room_total_rooms = 0
    if tab == 'rooms':
        from collections import OrderedDict
        rq = Room.query.filter(
            Room.is_active == True,
            Room.class_name.isnot(None),
            Room.class_name != ''
        )
        if sel_grade:
            rq = rq.filter_by(grade=sel_grade)
        rooms = rq.order_by(
            Room.grade, Room.gender, Room.class_name,
            Room.building, Room.room_number
        ).all()
        room_tree = OrderedDict()
        for room in rooms:
            g = room.grade or ''
            gender = room.gender or ''
            cn = room.class_name or ''
            if g not in room_tree:
                room_tree[g] = OrderedDict()
            if gender not in room_tree[g]:
                room_tree[g][gender] = OrderedDict()
            if cn not in room_tree[g][gender]:
                room_tree[g][gender][cn] = []
            room_tree[g][gender][cn].append(room)
        for g in room_tree:
            for gender in room_tree[g]:
                for cn in room_tree[g][gender]:
                    cnt = Student.query.filter_by(
                        grade=g, class_name=cn, gender=gender,
                        boarding_type='住校'
                    ).count()
                    room_class_totals[(g, cn, gender)] = cnt
        room_total_rooms = len(rooms)
        room_total_beds = sum(r.capacity for r in rooms)
        room_total_boarders = sum(room_class_totals.values())

    return render_template('dormitory/statistics/overview.html',
                           tab=tab,
                           sel_grade=sel_grade,
                           grade_options=grade_options,
                           per_class_stats=per_class_stats,
                           per_grade_stats=per_grade_stats,
                           school_stats=school_stats,
                           dorm_stats=_dorm_stats(),
                           scope_type=scope_type,
                           room_tree=room_tree,
                           room_class_totals=room_class_totals,
                           room_total_rooms=room_total_rooms,
                           room_total_beds=room_total_beds,
                           room_total_boarders=room_total_boarders)


# ---- 宿舍历史查询 ----

@bp.route('/history')
@login_required
@perm_required('statistics.view')
def dormitory_history():
    """宿舍历史查询"""
    import sqlite3 as _sql
    import os as _os
    from config import BASE_DIR
    history_db_path = _os.path.join(BASE_DIR, 'data', 'history.db')

    graduated_grades = []
    if _os.path.exists(history_db_path):
        try:
            conn = _sql.connect(history_db_path)
            grades = conn.execute(
                'SELECT DISTINCT graduated_grade FROM graduated_rooms ORDER BY graduated_grade'
            ).fetchall()
            graduated_grades = [g[0] for g in grades if g[0]]
            conn.close()
        except Exception:
            pass

    search_grade = request.args.get('grade', '')
    search_room = request.args.get('room_number', '').strip()
    search_building = request.args.get('building', '').strip()
    search_class = request.args.get('class_name', '').strip()

    rooms = []
    beds = []
    if search_grade and _os.path.exists(history_db_path):
        try:
            conn = _sql.connect(history_db_path)
            conn.row_factory = _sql.Row

            room_where = ['graduated_grade = ?']
            room_params = [search_grade]
            if search_room:
                room_where.append('room_number LIKE ?')
                room_params.append(f'%{search_room}%')
            if search_building:
                room_where.append('building LIKE ?')
                room_params.append(f'%{search_building}%')
            if search_class:
                room_where.append('class_name LIKE ?')
                room_params.append(f'%{search_class}%')

            room_sql = f'SELECT * FROM graduated_rooms WHERE {" AND ".join(room_where)} ORDER BY building, room_number LIMIT 200'
            room_rows = conn.execute(room_sql, room_params).fetchall()
            rooms = [dict(r) for r in room_rows]

            room_ids = [r['original_room_id'] for r in rooms if r['original_room_id']]
            if room_ids:
                ph = ','.join('?' * len(room_ids))
                bed_rows = conn.execute(
                    f'SELECT * FROM graduated_beds WHERE original_room_id IN ({ph}) ORDER BY original_room_id, bed_number',
                    room_ids
                ).fetchall()
                beds = [dict(b) for b in bed_rows]

            conn.close()
        except Exception as e:
            flash(f'查询失败：{str(e)}', 'danger')

    return render_template('dormitory/statistics/history.html',
                           graduated_grades=graduated_grades,
                           search_grade=search_grade,
                           search_room=search_room,
                           search_building=search_building,
                           search_class=search_class,
                           rooms=rooms,
                           beds=beds)


@bp.route('/alumni')
@login_required
@perm_required('statistics.view')
def alumni():
    """往届生宿舍查询"""
    import sqlite3 as _sql
    import os as _os
    from config import BASE_DIR
    history_db_path = _os.path.join(BASE_DIR, 'data', 'history.db')

    graduated_grades = []
    if _os.path.exists(history_db_path):
        try:
            conn = _sql.connect(history_db_path)
            grades = conn.execute(
                'SELECT DISTINCT graduated_grade FROM graduated_rooms ORDER BY graduated_grade'
            ).fetchall()
            graduated_grades = [g[0] for g in grades if g[0]]
            conn.close()
        except Exception:
            pass

    search_grade = request.args.get('grade', '')
    search_room = request.args.get('room_number', '').strip()
    search_building = request.args.get('building', '').strip()

    rooms = []
    beds = []
    if search_grade and _os.path.exists(history_db_path):
        try:
            conn = _sql.connect(history_db_path)
            conn.row_factory = _sql.Row

            room_where = ['graduated_grade = ?']
            room_params = [search_grade]
            if search_room:
                room_where.append('room_number LIKE ?')
                room_params.append(f'%{search_room}%')
            if search_building:
                room_where.append('building LIKE ?')
                room_params.append(f'%{search_building}%')

            room_sql = f'SELECT * FROM graduated_rooms WHERE {" AND ".join(room_where)} ORDER BY building, room_number LIMIT 200'
            room_rows = conn.execute(room_sql, room_params).fetchall()
            rooms = [dict(r) for r in room_rows]

            room_ids = [r['original_room_id'] for r in rooms if r['original_room_id']]
            if room_ids:
                ph = ','.join('?' * len(room_ids))
                bed_rows = conn.execute(
                    f'SELECT * FROM graduated_beds WHERE original_room_id IN ({ph}) ORDER BY original_room_id, bed_number',
                    room_ids
                ).fetchall()
                beds = [dict(b) for b in bed_rows]

            conn.close()
        except Exception as e:
            flash(f'查询失败：{str(e)}', 'danger')

    return render_template('dormitory/statistics/alumni.html',
                           graduated_grades=graduated_grades,
                           search_grade=search_grade,
                           search_room=search_room,
                           search_building=search_building,
                           rooms=rooms,
                           beds=beds)


# ---- 宿舍数据导入（宿管专用）----

_import_errors = {}


def _save_import_errors(errors):
    key = uuid.uuid4().hex
    _import_errors[key] = (errors, time.time() + 3600)
    return key


def _clean_expired_errors():
    now = time.time()
    expired = [k for k, v in _import_errors.items() if v[1] < now]
    for k in expired:
        del _import_errors[k]


@bp.route('/download-dormitory-template')
@perm_required('dormitory.import')
def download_dormitory_template():
    """下载宿舍数据导入模板"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '宿舍数据导入'

    headers = ['学号', '姓名', '住校/走读', '出门权限', '课本', '班主任备注']

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    required_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    required_cols = {0, 1}

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = required_fill if (col_idx - 1) in required_cols else header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    widths = [15, 10, 12, 12, 10, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.append(['20251001', '张三', '住校', '', '', ''])

    ws.merge_cells('A2:F2')
    instr_cell = ws.cell(row=2, column=1,
        value='说明：橙色列必填 | 住校/走读填住校/男走读/女走读/离校 | 出门权限填晚走读/午晚走读/艺术生 | 第3行起填数据，删除本行和示例')
    instr_cell.font = Font(color='FF0000', bold=True, size=10)
    instr_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 30

    for col in range(1, len(headers) + 1):
        ws.cell(row=3, column=col).border = thin_border
        ws.cell(row=3, column=col).alignment = Alignment(horizontal='center')

    for row in ws.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.border = thin_border

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A1:F{ws.max_row}'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name='宿舍数据导入模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/import-dormitory', methods=['POST'])
@perm_required('dormitory.import')
def import_dormitory():
    """批量导入宿舍数据（学号+姓名双重匹配）"""
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('请上传 .xlsx 格式的Excel文件', 'danger')
        return redirect(url_for('statistics.index'))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        header_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                header_map[str(val).strip()] = col

        field_mapping = {
            '学号': 'student_number', '姓名': 'name',
            '住校/走读': 'boarding_type', '出门权限': 'day_student_type',
            '课本': 'textbook', '班主任备注': 'teacher_notes',
        }

        col_map = {}
        for excel_name, model_field in field_mapping.items():
            if excel_name in header_map and model_field not in col_map:
                col_map[model_field] = header_map[excel_name]

        required_fields = {'name': '姓名', 'student_number': '学号'}
        missing = [v for k, v in required_fields.items() if k not in col_map]
        if missing:
            flash(f'Excel缺少必填列：{", ".join(missing)}', 'danger')
            return redirect(url_for('statistics.index'))

        all_students = {}
        for s in Student.query.all():
            all_students[s.student_number] = s

        updated_count = 0
        errors = []

        for row_idx in range(2, ws.max_row + 1):
            try:
                row_data = {}
                for field, col in col_map.items():
                    cell = ws.cell(row=row_idx, column=col)
                    v = cell.value
                    if v is not None:
                        v = str(v).strip()
                    row_data[field] = v

                student_number = row_data.get('student_number', '')
                name = row_data.get('name', '')

                if not student_number:
                    continue

                if student_number not in all_students:
                    errors.append(f'第{row_idx}行（{name}）：学号 {student_number} 在基础数据中不存在')
                    continue

                student = all_students[student_number]
                if student.name != name:
                    errors.append(f'第{row_idx}行（{name}）：学号 {student_number} 对应学生姓名为「{student.name}」，与Excel中「{name}」不匹配')
                    continue

                acc = StudentAccommodation.query.filter_by(student_id=student.id).first()
                if not acc:
                    acc = StudentAccommodation(student_id=student.id)
                    db.session.add(acc)
                
                acc.boarding_type = row_data.get('boarding_type') or acc.boarding_type
                acc.day_student_type = row_data.get('day_student_type') or acc.day_student_type
                acc.textbook = row_data.get('textbook') or acc.textbook
                acc.teacher_notes = row_data.get('teacher_notes') or acc.teacher_notes
                updated_count += 1

            except Exception as e:
                errors.append(f'第{row_idx}行处理异常：{str(e)}')
                continue

        if updated_count > 0:
            db.session.commit()
            log_operation(current_user, '导入', '宿舍数据', None, f'批量更新 {updated_count} 名学生宿舍信息')

        if errors:
            ek = _save_import_errors(errors)
            flash(f'成功更新 {updated_count} 名学生宿舍信息，{len(errors)} 条失败', 'warning')
            flash(Markup(f'<a href="/statistics/download-errors/{ek}" class="btn btn-sm btn-outline-danger">下载错误日志 ({len(errors)}条)</a>'), 'warning')
        elif updated_count > 0:
            flash(f'成功更新 {updated_count} 名学生宿舍信息', 'success')
        else:
            flash('Excel中没有有效的宿舍数据', 'warning')

    except Exception as e:
        db.session.rollback()
        flash(f'导入失败：{str(e)}', 'danger')

    return redirect(url_for('statistics.index'))


@bp.route('/download-errors/<key>')
@login_required
def download_import_errors(key):
    """下载导入错误日志"""
    _clean_expired_errors()
    data = _import_errors.pop(key, None)
    if not data:
        flash('错误日志已过期或不存在', 'warning')
        return redirect(url_for('statistics.index'))
    errors, _ = data
    content = '\r\n'.join(errors)
    from io import BytesIO
    buf = BytesIO()
    buf.write(content.encode('utf-8-sig'))
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'导入错误日志_{key[:8]}.txt',
                     mimetype='text/plain; charset=utf-8')


