# StuLink v1.6.1 2026-07-09
# Copyright (c) 2026 zkxxzf. Apache License 2.0
import secrets
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from app.extensions import db
from app.models import User, PermissionGroup
from app.forms.user_forms import UserForm
from app.utils.decorators import perm_required
from app.utils.helpers import log_operation
import io

bp = Blueprint('users', __name__, url_prefix='/users')


def _generate_password():
    """生成随机安全密码（10位字母数字）"""
    import string
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(10))


@bp.route('/')
@perm_required('system.users')
def list_users():
    users = User.query.order_by(User.role, User.username).all()
    return render_template('system/users/list.html', users=users)


@bp.route('/create', methods=['GET', 'POST'])
@perm_required('system.users')
def create():
    form = UserForm()
    if form.validate_on_submit():
        if User.query.filter_by(username=form.username.data).first():
            flash('用户名已存在', 'danger')
            return render_template('system/users/form.html', form=form, title='新建用户')

        # 校验：班主任最多3人/班，且年级班级必填
        role = form.role.data
        grade = form.grade.data or None
        class_name = form.class_name.data or None

        if role == 'homeroom_teacher':
            if not grade or not class_name:
                flash('班主任必须指定年级和班级', 'danger')
                return render_template('system/users/form.html', form=form, title='新建用户')
            existing_count = User.query.filter_by(
                role='homeroom_teacher', grade=grade, class_name=class_name, is_active=True
            ).count()
            if existing_count >= 3:
                flash(f'{grade}{class_name} 已有 {existing_count} 名班主任，最多设置3名', 'danger')
                return render_template('system/users/form.html', form=form, title='新建用户')

        if role == 'grade_leader' and not grade:
            flash('年级长必须指定管理年级', 'danger')
            return render_template('system/users/form.html', form=form, title='新建用户')

        user = User(
            username=form.username.data,
            real_name=form.real_name.data,
            role=role,
            grade=grade,
            class_name=class_name,
            permission_group_id=form.permission_group_id.data if form.permission_group_id.data else None,
            must_change_pwd=True,
        )
        generated_pwd = form.password.data or _generate_password()
        user.set_password(generated_pwd)
        db.session.add(user)
        db.session.commit()
        log_operation(current_user, '创建', '用户', user.id, f'{user.real_name} ({user.role_display})')
        pwd_hint = f'，初始密码：{generated_pwd}' if not form.password.data else ''
        flash(f'用户 {user.real_name} 已创建{pwd_hint}', 'success')
        return redirect(url_for('users.list_users'))
    return render_template('system/users/form.html', form=form, title='新建用户')


@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@perm_required('system.users')
def edit(id):
    user = User.query.get_or_404(id)
    form = UserForm(obj=user)
    if form.validate_on_submit():
        existing = User.query.filter(
            User.username == form.username.data, User.id != user.id).first()
        if existing:
            flash('用户名已存在', 'danger')
            return render_template('system/users/form.html', form=form, title='编辑用户')

        # 校验
        role = form.role.data
        grade = form.grade.data or None
        class_name = form.class_name.data or None

        if role == 'homeroom_teacher':
            if not grade or not class_name:
                flash('班主任必须指定年级和班级', 'danger')
                return render_template('system/users/form.html', form=form, title='编辑用户')
            existing_count = User.query.filter_by(
                role='homeroom_teacher', grade=grade, class_name=class_name, is_active=True
            ).filter(User.id != user.id).count()
            if existing_count >= 3:
                flash(f'{grade}{class_name} 已有 {existing_count} 名班主任，最多设置3名', 'danger')
                return render_template('system/users/form.html', form=form, title='编辑用户')

        if role == 'grade_leader' and not grade:
            flash('年级长必须指定管理年级', 'danger')
            return render_template('system/users/form.html', form=form, title='编辑用户')

        user.username = form.username.data
        user.real_name = form.real_name.data
        user.role = role
        user.grade = grade
        user.class_name = class_name
        user.permission_group_id = form.permission_group_id.data if form.permission_group_id.data else None
        if form.password.data:
            user.set_password(form.password.data)
        db.session.commit()
        log_operation(current_user, '更新', '用户', user.id, f'{user.real_name} 信息已更新')
        flash('用户信息已更新', 'success')
        return redirect(url_for('users.list_users'))
    return render_template('system/users/form.html', form=form, title='编辑用户')


@bp.route('/<int:id>/toggle', methods=['POST'])
@perm_required('system.users')
def toggle(id):
    user = User.query.get_or_404(id)
    user.is_active = not user.is_active
    db.session.commit()
    status = '启用' if user.is_active else '禁用'
    log_operation(current_user, '更新', '用户', user.id, f'{user.real_name} {status}')
    flash(f'用户 {user.real_name} 已{status}', 'success')
    return redirect(url_for('users.list_users'))


@bp.route('/<int:id>/reset-password', methods=['POST'])
@perm_required('system.users')
def reset_password(id):
    user = User.query.get_or_404(id)
    user.set_password(_generate_password())
    user.must_change_pwd = True
    db.session.commit()
    log_operation(current_user, '更新', '用户', user.id, f'{user.real_name} 密码已重置')
    flash(f'{user.real_name} 的密码已重置', 'success')
    return redirect(url_for('users.list_users'))


@bp.route('/download-teacher-template')
@perm_required('system.users')
def download_teacher_template():
    """下载教师批量导入模板"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '教师导入模板'

    headers = ['手机号', '真实姓名', '权限组']
    header_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14

    # example row
    ws.append(['13800138000', '张老师', '管理员'])

    # instruction row
    ws.merge_cells('A4:C4')
    instr = ws.cell(row=4, column=1, value='说明：手机号即登录名，默认密码=手机号，首次登录需改密 | 权限组填写系统已有权限组名称 | 第 2 行起填数据，删除本行和示例')
    instr.font = Font(color='FF0000', bold=True, size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='教师导入模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/import-teachers', methods=['POST'])
@perm_required('system.users')
def import_teachers():
    """批量导入教师：手机号=登录名，默认密码=手机号，首次登录强制改密"""
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('请上传 .xlsx 格式的Excel文件', 'danger')
        return redirect(url_for('users.list_users'))

    try:
        import openpyxl.worksheet.datavalidation as dv
        _orig_init = dv.DataValidation.__init__
        def _patched_init(self, *a, **kw):
            kw.pop('id', None)
            _orig_init(self, *a, **kw)
        dv.DataValidation.__init__ = _patched_init

        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        header_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                header_map[str(val).strip()] = col

        phone_col = header_map.get('手机号')
        name_col = header_map.get('真实姓名')
        group_col = header_map.get('权限组')

        if not phone_col or not name_col or not group_col:
            flash('Excel缺少必填列：手机号、真实姓名、权限组', 'danger')
            return redirect(url_for('users.list_users'))

        # load existing phones and permission groups
        existing_phones = set(r[0] for r in db.session.query(User.username).all() if r[0])
        pg_map = {pg.name: pg for pg in PermissionGroup.query.all()}

        created = 0
        errors = []

        for row_idx in range(2, ws.max_row + 1):
            phone = str(ws.cell(row=row_idx, column=phone_col).value or '').strip()
            name = str(ws.cell(row=row_idx, column=name_col).value or '').strip()
            group_name = str(ws.cell(row=row_idx, column=group_col).value or '').strip()

            if not phone or not name or not group_name:
                continue

            if phone in existing_phones:
                errors.append(f'第{row_idx}行（{name}）：手机号{phone}已存在')
                continue

            pg = pg_map.get(group_name)
            if not pg:
                errors.append(f'第{row_idx}行（{name}）：权限组"{group_name}"不存在')
                continue

            user = User(username=phone, real_name=name, role='teacher',
                        permission_group_id=pg.id,
                        must_change_pwd=True, is_active=True)
            user.set_password(phone)
            db.session.add(user)
            existing_phones.add(phone)
            created += 1

        db.session.commit()

        if created:
            log_operation(current_user, '导入', '教师', None, f'批量导入 {created} 名教师')
        msg = f'成功导入 {created} 名教师'
        if errors:
            msg += f'，{len(errors)} 条失败：' + '；'.join(errors[:20])
            if len(errors) > 20:
                msg += f'...还有{len(errors)-20}条'
            flash(msg, 'warning')
        else:
            flash(msg, 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'导入失败：{str(e)}', 'danger')

    return redirect(url_for('users.list_users'))


