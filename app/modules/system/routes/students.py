# StuLink v1.6.1 2026-07-09
# Copyright (c) 2026 zkxxzf. Apache License 2.0
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from markupsafe import Markup
from flask_login import login_required, current_user
from app.extensions import db
from app.models import Student, BedAssignment, StudentAccommodation
from app.utils.crypto import encrypt as _encrypt_id
from app.forms.student_forms import StudentForm
from app.utils.decorators import perm_required
from app.utils.helpers import get_dict_values, log_operation, get_graduated_grades, write_change_log
import io
import uuid
import time

bp = Blueprint('students', __name__, url_prefix='/students')


@bp.route('/export', methods=['GET', 'POST'])
@login_required
def export_students():
    from app.utils.export_helpers import do_export_students
    if request.method == 'POST':
        return do_export_students(request.form)
    return do_export_students(request.args)

# 导入错误日志暂存（key=uuid, value=(errors_list, expiry_time)）
_import_errors = {}


def _save_import_errors(errors):
    """保存导入错误，返回uuid key"""
    key = uuid.uuid4().hex
    _import_errors[key] = (errors, time.time() + 3600)  # 1小时过期
    return key


def _clean_expired_errors():
    now = time.time()
    expired = [k for k, v in _import_errors.items() if v[1] < now]
    for k in expired:
        del _import_errors[k]


@bp.route('/')
@login_required
def list_students():
    query = Student.query
    # 排除已毕业年级
    graduated = get_graduated_grades()
    if graduated:
        query = query.filter(~Student.grade.in_(graduated))
    # 权限范围限制
    if current_user.role == 'homeroom_teacher':
        query = query.filter_by(grade=current_user.grade, class_name=current_user.class_name)
    elif current_user.role == 'grade_leader':
        query = query.filter_by(grade=current_user.grade)

    # 筛选参数
    filter_gender = request.args.get('gender', '')
    filter_grade = request.args.get('grade', '')
    filter_class = request.args.get('class_name', '')
    filter_subject = request.args.get('subject_selection', '')
    filter_school = request.args.get('graduation_school', '').strip()
    filter_enrollment = request.args.get('enrollment_status', '')
    filter_name = request.args.get('name', '').strip()
    filter_student_number = request.args.get('student_number', '').strip()

    if filter_gender:
        query = query.filter_by(gender=filter_gender)
    if filter_grade:
        query = query.filter_by(grade=filter_grade)
    if filter_class:
        query = query.filter_by(class_name=filter_class)
    if filter_subject:
        query = query.filter_by(subject_selection=filter_subject)
    if filter_school:
        query = query.filter(Student.graduation_school.contains(filter_school))
    if filter_enrollment:
        query = query.filter_by(enrollment_status=filter_enrollment)
    if filter_name:
        query = query.filter(Student.name.contains(filter_name))
    if filter_student_number:
        query = query.filter(Student.student_number.contains(filter_student_number))
    
    # 分页参数
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    pagination = query.order_by(Student.grade, Student.class_name, Student.student_number).paginate(
        page=page, per_page=per_page, error_out=False
    )
    students = pagination.items
    
    grades = get_dict_values('grade')
    grades = [g for g in grades if g not in graduated]
    classes = get_dict_values('class')
    subjects = get_dict_values('subject')
    boarding_types = get_dict_values('boarding_type')
    day_student_types = get_dict_values('day_student_type')
    enrollment_statuses = get_dict_values('enrollment_status')
    # 毕业学校列表
    school_list = [s[0] for s in db.session.query(Student.graduation_school)
                   .filter(Student.graduation_school.isnot(None), Student.graduation_school != '')
                   .distinct().order_by(Student.graduation_school).all()
                   if s[0]]
    return render_template('system/students/list.html', 
                         students=students, 
                         grades=grades, 
                         classes=classes,
                         subjects=subjects,
                         boarding_types=boarding_types,
                         day_student_types=day_student_types,
                         enrollment_statuses=enrollment_statuses,
                         pagination=pagination,
                         filter_gender=filter_gender,
                         filter_grade=filter_grade,
                         filter_class=filter_class,
                         filter_subject=filter_subject,
                         filter_school=filter_school,
                         filter_enrollment=filter_enrollment,
                         school_list=school_list)



@bp.route('/create', methods=['GET', 'POST'])
@perm_required('students.edit')
def create():
    form = StudentForm()
    if form.validate_on_submit():
        # 1. 验证身份证号格式
        id_card = form.id_card_number.data
        if id_card and not _validate_id_card(id_card):
            flash('身份证号格式不正确', 'danger')
            return render_template('system/students/form.html', form=form, title='新增学生')
        
        # 2. 检查身份证号是否重复
        if id_card and Student.query.filter_by(_id_card_encrypted=_encrypt_id(id_card)).first():
            flash('该身份证号已存在', 'danger')
            return render_template('system/students/form.html', form=form, title='新增学生')
        
        # 3. 检查学号唯一性（学号不为空时）
        if form.student_number.data:
            if Student.query.filter_by(student_number=form.student_number.data).first():
                flash('该学号已存在', 'danger')
                return render_template('system/students/form.html', form=form, title='新增学生')
        
        student = Student()
        _populate_student(student, form)
        db.session.add(student)
        db.session.commit()
        log_operation(current_user, '创建', '学生', student.id, f'{student.name} {student.grade}{student.class_name}')
        flash('学生信息已添加', 'success')
        return redirect(url_for('students.list_students'))
    # 班主任默认填入自己管理的年级班级
    if current_user.role == 'homeroom_teacher':
        form.grade.data = form.grade.data or current_user.grade
        form.class_name.data = form.class_name.data or current_user.class_name
    return render_template('system/students/form.html', form=form, title='新增学生')


@bp.route('/<int:id>')
@login_required
def detail(id):
    student = Student.query.get_or_404(id)
    # 权限控制：宿管教师不能查看学生隐私数据
    # 仅允许：管理员、班主任(仅本班)、年级长(仅本年级)
    if current_user.role == 'dorm_manager':
        flash('无权查看学生详细信息', 'danger')
        return redirect(url_for('students.list_students'))
    if current_user.role == 'homeroom_teacher':
        if student.grade != current_user.grade or student.class_name != current_user.class_name:
            flash('无权查看该学生信息', 'danger')
            return redirect(url_for('students.list_students'))
    elif current_user.role == 'grade_leader':
        if student.grade != current_user.grade:
            flash('无权查看该年级学生信息', 'danger')
            return redirect(url_for('students.list_students'))
    elif current_user.role not in ('admin', 'school_viewer'):
        flash('无权查看学生详细信息', 'danger')
        return redirect(url_for('students.list_students'))
    # 查询变迁记录
    change_logs = []
    try:
        import sqlite3, os
        from config import BASE_DIR
        history_path = os.path.join(BASE_DIR, 'data', 'history.db')
        if os.path.exists(history_path):
            conn = sqlite3.connect(history_path)
            conn.row_factory = sqlite3.Row
            conn.execute('''
                CREATE TABLE IF NOT EXISTS student_change_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    student_id INTEGER NOT NULL,
                    student_number TEXT,
                    student_name TEXT NOT NULL,
                    change_type TEXT NOT NULL,
                    old_value TEXT,
                    new_value TEXT,
                    detail TEXT,
                    operator TEXT,
                    changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            rows = conn.execute(
                'SELECT * FROM student_change_log WHERE student_id=? ORDER BY changed_at DESC',
                (student.id,)
            ).fetchall()
            change_logs = [dict(r) for r in rows]
            conn.close()
    except Exception:
        pass
    return render_template('system/students/detail.html', student=student, change_logs=change_logs)


@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@perm_required('students.edit')
def edit(id):
    student = Student.query.get_or_404(id)
    # 权限控制：检查用户是否有权限编辑该学生
    if current_user.role == 'homeroom_teacher':
        if student.grade != current_user.grade or student.class_name != current_user.class_name:
            flash('无权编辑该学生', 'danger')
            return redirect(url_for('students.list_students'))
    elif current_user.role == 'grade_leader':
        if student.grade != current_user.grade:
            flash('无权编辑该年级学生', 'danger')
            return redirect(url_for('students.list_students'))
    elif current_user.role not in ('admin',):
        flash('无权编辑学生', 'danger')
        return redirect(url_for('students.list_students'))
    form = StudentForm(obj=student)
    if form.validate_on_submit():
        # 1. 验证身份证号格式
        id_card = form.id_card_number.data
        if id_card and not _validate_id_card(id_card):
            flash('身份证号格式不正确', 'danger')
            return render_template('system/students/form.html', form=form, title='编辑学生')
        
        # 2. 检查身份证号是否重复（排除自身）
        if id_card:
            existing_id = Student.query.filter(
                Student._id_card_encrypted == _encrypt_id(id_card),
                Student.id != student.id
            ).first()
            if existing_id:
                flash('该身份证号已存在', 'danger')
                return render_template('system/students/form.html', form=form, title='编辑学生')
        
        # 3. 检查学号唯一性（排除自身，学号不为空时）
        if form.student_number.data:
            existing_sn = Student.query.filter(
                Student.student_number == form.student_number.data,
                Student.id != student.id
            ).first()
            if existing_sn:
                flash('该学号已存在', 'danger')
                return render_template('system/students/form.html', form=form, title='编辑学生')
        
        _populate_student_for_edit(student, form)
        db.session.commit()
        log_operation(current_user, '更新', '学生', student.id, f'{student.name} {student.grade}{student.class_name}')
        flash('学生信息已更新', 'success')
        return redirect(url_for('students.detail', id=student.id))
    return render_template('system/students/form.html', form=form, title='编辑学生')


@bp.route('/<int:id>/delete', methods=['POST'])
@perm_required('students.import')
def delete(id):
    student = Student.query.get_or_404(id)
    # 权限控制：检查用户是否有权限删除该学生
    if current_user.role == 'homeroom_teacher':
        if student.grade != current_user.grade or student.class_name != current_user.class_name:
            flash('无权删除该学生', 'danger')
            return redirect(url_for('students.list_students'))
    elif current_user.role == 'grade_leader':
        if student.grade != current_user.grade:
            flash('无权删除该年级学生', 'danger')
            return redirect(url_for('students.list_students'))
    elif current_user.role not in ('admin',):
        flash('无权删除学生', 'danger')
        return redirect(url_for('students.list_students'))
    # 先删除床位分配
    if student.bed_assignment:
        student.bed_assignment.student_id = None
    db.session.delete(student)
    db.session.commit()
    log_operation(current_user, '删除', '学生', id, f'{student.name} {student.grade}{student.class_name}')
    flash(f'学生 {student.name} 已删除', 'success')
    return redirect(url_for('students.list_students'))


@bp.route('/search')
@login_required
def search():
    query = Student.query
    # 排除已毕业年级
    graduated = get_graduated_grades()
    if graduated:
        query = query.filter(~Student.grade.in_(graduated))
    # 获取搜索条件
    name = request.args.get('name', '').strip()
    student_number = request.args.get('student_number', '').strip()
    grade = request.args.get('grade', '')
    class_name = request.args.get('class_name', '')
    gender = request.args.get('gender', '')
    boarding_type = request.args.get('boarding_type', '')
    day_student_type = request.args.get('day_student_type', '')
    subject_selection = request.args.get('subject_selection', '')
    room_number = request.args.get('room_number', '').strip()

    if name:
        query = query.filter(Student.name.contains(name))
    if student_number:
        query = query.filter(Student.student_number.contains(student_number))
    if grade:
        query = query.filter_by(grade=grade)
    if class_name:
        query = query.filter_by(class_name=class_name)
    if gender:
        query = query.filter_by(gender=gender)
    if boarding_type:
        acc_ids = [sa.student_id for sa in StudentAccommodation.query.filter_by(boarding_type=boarding_type).all()]
        if acc_ids:
            query = query.filter(Student.id.in_(acc_ids))
        else:
            query = query.filter(Student.id == -1)
    if day_student_type:
        acc_ids = [sa.student_id for sa in StudentAccommodation.query.filter_by(day_student_type=day_student_type).all()]
        if acc_ids:
            query = query.filter(Student.id.in_(acc_ids))
        else:
            query = query.filter(Student.id == -1)
    if subject_selection:
        query = query.filter_by(subject_selection=subject_selection)
    # 宿舍号筛选
    if room_number:
        from app.models import BedAssignment
        bed_sub = db.session.query(BedAssignment.student_id).join(BedAssignment.room).filter(
            Room.room_number.contains(room_number)
        ).filter(BedAssignment.student_id.isnot(None)).all()
        bed_ids = [b[0] for b in bed_sub if b[0]]
        if bed_ids:
            query = query.filter(Student.id.in_(bed_ids))
        else:
            query = query.filter(Student.id == -1)  # 无匹配

    # 分页参数
    page = request.args.get('page', 1, type=int)
    per_page = 50  # 每页显示50条
    
    # 使用分页查询
    pagination = query.order_by(Student.grade, Student.class_name, Student.student_number).paginate(
        page=page, per_page=per_page, error_out=False
    )
    students = pagination.items
    
    grades = get_dict_values('grade')
    grades = [g for g in grades if g not in graduated]
    classes = get_dict_values('class')
    boarding_types = get_dict_values('boarding_type')
    day_student_types = get_dict_values('day_student_type')
    subjects = get_dict_values('subject')
    buildings = get_dict_values('building')
    return render_template('system/students/search.html', 
                         students=students, 
                         grades=grades, 
                         classes=classes,
                         boarding_types=boarding_types, 
                         day_student_types=day_student_types, 
                         subjects=subjects,
                         buildings=buildings,
                         pagination=pagination)


@bp.route('/batch-edit-search', methods=['POST'])
@login_required
def batch_edit_search():
    """批量修改学生宿舍信息（搜索页）"""
    if not current_user.has_perm('dormitory.manage') and not current_user.has_perm('students.edit'):
        return jsonify({'success': False, 'message': '无权限'}), 403

    data = request.get_json() or {}
    ids = data.get('student_ids', [])
    if not ids:
        return jsonify({'success': False, 'message': '请选择学生'})

    boarding_type = data.get('boarding_type', '')
    day_student_type = data.get('day_student_type', '')
    building = data.get('building', '')
    room_number = data.get('room_number', '')
    bed_number = data.get('bed_number', '')

    # 班主任只能改自己班
    if current_user.role == 'homeroom_teacher':
        students = Student.query.filter(
            Student.id.in_(ids),
            Student.grade == current_user.grade,
            Student.class_name == current_user.class_name
        ).all()
    else:
        students = Student.query.filter(Student.id.in_(ids)).all()

    if not students:
        return jsonify({'success': False, 'message': '无匹配学生或无权限'})

    updated = 0
    # 批量修改床位
    if building and room_number and bed_number:
        try:
            bed_num = int(bed_number)
        except ValueError:
            return jsonify({'success': False, 'message': '床位号必须是数字'})
        # 查房间
        room = Room.query.filter_by(building=building, room_number=room_number, is_active=True).first()
        if not room:
            return jsonify({'success': False, 'message': f'{building} {room_number} 不存在'})
        # 验证床位有效性
        bed = BedAssignment.query.filter_by(room_id=room.id, bed_number=bed_num).first()
        if not bed:
            return jsonify({'success': False, 'message': f'{room_number} {bed_num}床 不存在（容量{room.capacity}人间）'})
        # 床位已被占
        if bed.student_id and bed.student_id not in [s.id for s in students]:
            occupant = Student.query.get(bed.student_id)
            return jsonify({'success': False, 'message': f'{building}{room_number} {bed_num}床已被{occupant.name if occupant else "其他学生"}占用'})

    for s in students:
        if boarding_type or day_student_type:
            acc = StudentAccommodation.query.filter_by(student_id=s.id).first()
            if not acc:
                acc = StudentAccommodation(student_id=s.id)
                db.session.add(acc)
            if boarding_type:
                acc.boarding_type = boarding_type
            if day_student_type:
                acc.day_student_type = day_student_type
        # 分配床位
        if building and room_number and bed_number:
            room = Room.query.filter_by(building=building, room_number=room_number, is_active=True).first()
            if room:
                bed = BedAssignment.query.filter_by(room_id=room.id, bed_number=int(bed_number)).first()
                if bed:
                    # 先清空旧床位
                    old_bed = BedAssignment.query.filter_by(student_id=s.id).first()
                    if old_bed:
                        old_bed.student_id = None
                    bed.student_id = s.id
                    # 同步更新房间班级信息
                    if not room.grade or not room.class_name:
                        room.grade = s.grade
                        room.class_name = s.class_name
        updated += 1

    db.session.commit()
    log_operation(current_user, '批量修改', '学生', None, f'搜索页批量修改 {updated} 人')
    return jsonify({'success': True, 'message': f'成功修改 {updated} 名学生'})


def _populate_student(student, form):
    """从表单填充学生对象（新增时使用）"""
    for field_name in ['name', 'gender', 'student_number', 'id_card_number', 'ethnicity',
                       'phone1', 'phone2', 'grade', 'class_name',
                       'subject_selection', 'enrollment_status', 'enrollment_notes',
                       'graduation_school_code', 'graduation_school']:
        setattr(student, field_name, getattr(form, field_name).data or None)


def _populate_student_for_edit(student, form):
    """从表单填充学生对象（编辑时使用）"""
    readonly_fields = ['name', 'gender', 'student_number', 'id_card_number', 'ethnicity',
                       'graduation_school_code', 'graduation_school']
    never_update_fields = ['grade', 'class_name']
    editable_fields = ['phone1', 'phone2', 'subject_selection', 'enrollment_status', 'enrollment_notes']
    
    for field_name in readonly_fields:
        if not getattr(student, field_name) and getattr(form, field_name).data:
            setattr(student, field_name, getattr(form, field_name).data or None)
    
    for field_name in editable_fields:
        setattr(student, field_name, getattr(form, field_name).data or None)


def _validate_id_card(id_card):
    """验证身份证号是否有效（18位）"""
    if not id_card:
        return False
    id_card = str(id_card).strip()
    if len(id_card) != 18:
        return False
    if not id_card[:17].isdigit():
        return False
    if not (id_card[17].isdigit() or id_card[17].upper() == 'X'):
        return False
    weights = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
    check_codes = '10X98765432'
    total = sum(int(id_card[i]) * weights[i] for i in range(17))
    if check_codes[total % 11] != id_card[17].upper():
        return False
    return True


@bp.route('/download-template')
@perm_required('students.import')
def download_template():
    """下载导入模板"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '学生导入模板'

    # 基础数据模板（不含宿舍相关字段，宿舍字段由宿管独立管理）
    headers = ['学号', '姓名', '性别', '民族', '身份证号', '学籍情况', '学籍备注',
               '年级', '班级', '联系方式 1', '联系方式 2',
               '毕业学校代码', '毕业学校',
               '选科']

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    required_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    required_cols = {0, 1, 2, 3, 4, 7, 8}  # 学号/姓名/性别/民族/身份证号/年级/班级

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = required_fill if (col_idx - 1) in required_cols else header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    widths = [15, 10, 8, 8, 22, 10, 20, 10, 8, 15, 15, 12, 25, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.append(['20251001', '张三', '男', '汉族', '110101200801011234', '借读', '',
               '2025级', '01班', '13800138000', '',
               '0440', '郑州市管城回族区第二中学',
               '史政地'])

    # 第 2 行：填写说明（合并单元格）
    ws.merge_cells('A2:N2')
    instr_cell = ws.cell(row=2, column=1, value='说明：橙色列必填 | 性别男/女 | 身份证 18 位 | 年级格式如"2025级" | 班级格式如"01班" | 毕业学校代码与名称关联 | 第 3 行起填数据，删除本行和示例')
    instr_cell.font = Font(color='FF0000', bold=True, size=10)
    instr_cell.alignment = Alignment(horizontal='left')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='学生导入模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@bp.route('/import', methods=['POST'])
@perm_required('students.import')
def import_students():
    """批量导入学生 或 更新已有学生"""
    file = request.files.get('file')
    update_mode = request.form.get('update_mode') == '1'
    
    if not file:
        flash('请选择要上传的Excel文件', 'danger')
        return redirect(url_for('students.list_students'))
    
    if not file.filename:
        flash('请选择要上传的Excel文件', 'danger')
        return redirect(url_for('students.list_students'))
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash(f'文件格式不正确，请上传 .xlsx 格式的Excel文件（当前文件：{file.filename}）', 'danger')
        return redirect(url_for('students.list_students'))

    try:
        import openpyxl.worksheet.datavalidation as dv
        _orig_init = dv.DataValidation.__init__
        def _patched_init(self, *a, **kw):
            kw.pop('id', None)
            _orig_init(self, *a, **kw)
        dv.DataValidation.__init__ = _patched_init

        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # 读取表头映射
        header_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                header_map[str(val).strip()] = col
        
        if not header_map:
            flash('Excel文件表头为空，请确保第一行为表头', 'danger')
            return redirect(url_for('students.list_students'))

        # Excel列名 → Student模型字段（仅基础数据）
        # 支持多种表头写法（兼容模板下载的表头）
        field_mapping = {
            '姓名': 'name', '性别': 'gender', '学号': 'student_number',
            '身份证号': 'id_card_number', '年级': 'grade',
            '班级': 'class_name', '新班级': 'class_name',
            '民族': 'ethnicity', '联系方式1': 'phone1', '联系方式2': 'phone2',
            '联系方式 1': 'phone1', '联系方式 2': 'phone2',
            '联系电话1': 'phone1', '联系电话2': 'phone2',
            '电话1': 'phone1', '电话2': 'phone2',
            '选科': 'subject_selection', '学籍情况': 'enrollment_status',
            '学籍备注': 'enrollment_notes',
            '毕业学校代码': 'graduation_school_code', '毕业学校': 'graduation_school',
        }

        col_map = {}
        for excel_name, model_field in field_mapping.items():
            if excel_name in header_map and model_field not in col_map:
                col_map[model_field] = header_map[excel_name]

        # 检查必填列
        if update_mode:
            required_fields = {'name': '姓名', 'student_number': '学号'}
        else:
            required_fields = {'name': '姓名', 'gender': '性别',
                               'id_card_number': '身份证号', 'student_number': '学号',
                               'ethnicity': '民族', 'grade': '年级',
                               'class_name': '班级'}
        missing = [v for k, v in required_fields.items() if k not in col_map]
        if missing:
            flash(f'Excel缺少必填列：{", ".join(missing)}，请使用系统下载的导入模板', 'danger')
            return redirect(url_for('students.list_students'))

        # 已有数据: student_number -> (name, id)
        existing_snum_map = {}
        existing_ids = set()
        for s in db.session.query(Student.student_number, Student.name, Student.id, Student._id_card_encrypted).all():
            if s[0]:
                existing_snum_map[s[0]] = (s[1], s[2])
            if s[3]:
                existing_ids.add(s[3])

        errors = []
        students_to_add = []
        students_updated = 0
        conflicts = []
        seen_snums = set()
        seen_ids = set()

        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for model_field, col_idx in col_map.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data[model_field] = str(val).strip() if val is not None else ''

            name = row_data.get('name', '')
            if not name:
                continue

            row_errors = []
            row_label = f'第{row_idx}行'
            gender = row_data.get('gender', '')
            snum = row_data.get('student_number', '')
            id_card = row_data.get('id_card_number', '')

            if not update_mode:
                if not gender:
                    row_errors.append('性别为空')
                elif gender not in ('男', '女'):
                    row_errors.append(f'性别"{gender}"无效')
                if not id_card:
                    row_errors.append('身份证号为空')
            if not snum:
                row_errors.append('学号为空')

            if snum:
                if update_mode and snum in existing_snum_map:
                    pass
                elif update_mode:
                    row_errors.append(f'学号"{snum}"不存在，无法更新')
                elif snum in existing_snum_map:
                    row_errors.append(f'学号"{snum}"已存在')
                elif snum in seen_snums:
                    row_errors.append(f'学号"{snum}"文件内重复')
                else:
                    seen_snums.add(snum)

            if id_card and not update_mode:
                if not _validate_id_card(id_card):
                    row_errors.append(f'身份证号无效')
                elif _encrypt_id(id_card) in existing_ids:
                    row_errors.append(f'身份证号已存在')
                elif _encrypt_id(id_card) in seen_ids:
                    row_errors.append(f'身份证号文件内重复')
                else:
                    seen_ids.add(_encrypt_id(id_card))

            if row_errors:
                errors.append(f'{row_label}（{name}）：{"; ".join(row_errors)}')
                continue

            # update mode: match by (student_number, name), fill blanks only, report conflicts
            if update_mode and snum in existing_snum_map:
                db_name, db_id = existing_snum_map[snum]
                if name != db_name:
                    errors.append(f'{row_label}（{name}）：学号{snum}在库中姓名为"{db_name}"，姓名不匹配')
                    continue
                student = Student.query.get(db_id)
                if student:
                    row_conflicts = []
                    changed = False
                    for field in ['ethnicity','grade','class_name',
                                  'phone1','phone2','subject_selection',
                                  'enrollment_status','enrollment_notes',
                                  'graduation_school_code','graduation_school']:
                        new_val = row_data.get(field, '').strip()
                        if not new_val:
                            continue
                        old_val = (getattr(student, field) or '').strip()
                        if not old_val:
                            setattr(student, field, new_val)
                            changed = True
                        elif old_val != new_val:
                            row_conflicts.append(f'{field}：库中"{old_val}" vs 导入"{new_val}"')
                    acc_fields = ['boarding_type', 'day_student_type', 'textbook', 'teacher_notes']
                    acc = StudentAccommodation.query.filter_by(student_id=student.id).first()
                    for field in acc_fields:
                        new_val = row_data.get(field, '').strip()
                        if not new_val:
                            continue
                        old_val = (getattr(acc, field) or '').strip() if acc else ''
                        if not old_val:
                            if not acc:
                                acc = StudentAccommodation(student_id=student.id)
                                db.session.add(acc)
                            setattr(acc, field, new_val)
                            changed = True
                        elif old_val != new_val:
                            row_conflicts.append(f'{field}：库中"{old_val}" vs 导入"{new_val}"')
                    if gender and gender != student.gender:
                        if not student.gender:
                            student.gender = gender
                            changed = True
                        else:
                            row_conflicts.append(f'性别：库中"{student.gender}" vs 导入"{gender}"')
                    if id_card and id_card != student.id_card_number:
                        if not student.id_card_number:
                            student.id_card_number = id_card
                            changed = True
                        else:
                            row_conflicts.append('身份证号冲突')
                    if row_conflicts:
                        conflicts.append(f'{row_label}（{name}，学号{snum}）：' + '；'.join(row_conflicts))
                    if changed:
                        students_updated += 1
                continue

            student = Student(
                name=name, gender=gender, student_number=snum or None,
                id_card_number=id_card,
                grade=row_data.get('grade') or '',
                class_name=row_data.get('class_name') or '未分班',
                ethnicity=row_data.get('ethnicity') or '汉族',
                phone1=row_data.get('phone1') or None,
                phone2=row_data.get('phone2') or None,
                subject_selection=row_data.get('subject_selection') or None,
                enrollment_status=row_data.get('enrollment_status') or None,
                enrollment_notes=row_data.get('enrollment_notes') or None,
                graduation_school_code=row_data.get('graduation_school_code') or None,
                graduation_school=row_data.get('graduation_school') or None,
            )
            students_to_add.append((student, {
                'boarding_type': row_data.get('boarding_type') or '住校',
                'day_student_type': row_data.get('day_student_type') or None,
                'textbook': row_data.get('textbook') or None,
                'teacher_notes': row_data.get('teacher_notes') or None,
            }))

        # commit and report
        if students_updated:
            db.session.commit()
        if students_to_add:
            actual_students = [s[0] for s in students_to_add]
            db.session.add_all(actual_students)
            db.session.flush()
            for student, acc_data in students_to_add:
                acc_boarding_type = acc_data['boarding_type']
                acc_day_student_type = acc_data['day_student_type']
                acc_textbook = acc_data['textbook']
                acc_teacher_notes = acc_data['teacher_notes']
                if acc_boarding_type or acc_day_student_type or acc_textbook or acc_teacher_notes:
                    acc = StudentAccommodation(
                        student_id=student.id,
                        boarding_type=acc_boarding_type,
                        day_student_type=acc_day_student_type,
                        textbook=acc_textbook,
                        teacher_notes=acc_teacher_notes
                    )
                    db.session.add(acc)
            db.session.commit()

        if update_mode:
            parts = []
            if students_updated: parts.append(f'{students_updated}人已更新')
            if students_to_add: parts.append(f'{len(students_to_add)}人新增')
            if errors: parts.append(f'{len(errors)}条错误')
            if conflicts: parts.append(f'{len(conflicts)}处冲突')
            summary = '，'.join(parts) if parts else '无变更'
            log_operation(current_user, '增量更新', '学生', None, summary)
            msg = f'更新结果：{summary}'
            if conflicts:
                msg += '\n--- 字段冲突（Excel值与库中已有值不一致，未更新） ---\n' + '\n'.join(conflicts[:30])
                if len(conflicts) > 30:
                    msg += f'\n...还有{len(conflicts)-30}处冲突'
            if errors:
                ek = _save_import_errors(errors)
                flash(Markup(f'<a href="/students/download-errors/{ek}" class="btn btn-sm btn-outline-secondary">下载完整错误日志 ({len(errors)}条)</a>'), 'info')
            flash(msg, 'success' if not errors and not conflicts else 'warning')
        elif errors and students_to_add:
            # 部分导入：正确的行导入，错误的行报告
            actual_students = [s[0] for s in students_to_add]
            db.session.add_all(actual_students)
            db.session.flush()
            for student, acc_data in students_to_add:
                acc_boarding_type = acc_data['boarding_type']
                acc_day_student_type = acc_data['day_student_type']
                acc_textbook = acc_data['textbook']
                acc_teacher_notes = acc_data['teacher_notes']
                if acc_boarding_type or acc_day_student_type or acc_textbook or acc_teacher_notes:
                    acc = StudentAccommodation(
                        student_id=student.id,
                        boarding_type=acc_boarding_type,
                        day_student_type=acc_day_student_type,
                        textbook=acc_textbook,
                        teacher_notes=acc_teacher_notes
                    )
                    db.session.add(acc)
            db.session.commit()
            log_operation(current_user, '导入', '学生', None, f'部分导入 {len(students_to_add)} 名，{len(errors)} 条失败')
            error_summary = f'成功导入 {len(students_to_add)} 名学生，但有 {len(errors)} 条数据未导入'
            ek = _save_import_errors(errors)
            flash(error_summary, 'warning')
            flash(Markup(f'<a href="/students/download-errors/{ek}" class="btn btn-sm btn-outline-danger">下载完整错误日志 ({len(errors)}条)</a>'), 'warning')
        elif errors:
            # 全部错误，无有效数据
            error_summary = f'发现 {len(errors)} 条错误，未导入任何数据'
            ek = _save_import_errors(errors)
            flash(error_summary, 'danger')
            flash(Markup(f'<a href="/students/download-errors/{ek}" class="btn btn-sm btn-outline-danger">下载完整错误日志 ({len(errors)}条)</a>'), 'danger')
        elif not students_to_add:
            flash('Excel中没有有效的学生数据', 'warning')
        else:
            actual_students = [s[0] for s in students_to_add]
            db.session.add_all(actual_students)
            db.session.flush()
            for student, acc_data in students_to_add:
                acc_boarding_type = acc_data['boarding_type']
                acc_day_student_type = acc_data['day_student_type']
                acc_textbook = acc_data['textbook']
                acc_teacher_notes = acc_data['teacher_notes']
                if acc_boarding_type or acc_day_student_type or acc_textbook or acc_teacher_notes:
                    acc = StudentAccommodation(
                        student_id=student.id,
                        boarding_type=acc_boarding_type,
                        day_student_type=acc_day_student_type,
                        textbook=acc_textbook,
                        teacher_notes=acc_teacher_notes
                    )
                    db.session.add(acc)
            db.session.commit()
            log_operation(current_user, '导入', '学生', None, f'批量导入 {len(students_to_add)} 名学生')
            flash(f'成功导入 {len(students_to_add)} 名学生', 'success')

    except Exception as e:
        db.session.rollback()
        error_str = str(e)
        friendly_msg = '导入失败：'
        if 'UNIQUE constraint failed' in error_str:
            if 'student_accommodation.student_id' in error_str:
                friendly_msg += '检测到重复的学生住宿记录，请检查是否重复导入同一批学生数据，或使用"增量更新"模式导入'
            elif 'students.student_number' in error_str:
                friendly_msg += '学号重复，请检查Excel中是否有重复学号，或该学号已在系统中存在'
            elif 'students._id_card_encrypted' in error_str:
                friendly_msg += '身份证号重复，请检查Excel中是否有重复身份证号，或该身份证号已在系统中存在'
            else:
                friendly_msg += '数据重复，请检查Excel中是否有重复数据'
        elif 'student_accommodation' in error_str:
            friendly_msg += '住宿信息导入失败，请检查数据格式是否正确'
        else:
            friendly_msg += '系统处理异常，请联系管理员'
        flash(friendly_msg, 'danger')

    return redirect(url_for('students.list_students'))


@bp.route('/download-errors/<key>')
@login_required
def download_import_errors(key):
    """下载导入错误日志"""
    _clean_expired_errors()
    data = _import_errors.pop(key, None)
    if not data:
        flash('错误日志已过期或不存在', 'warning')
        return redirect(url_for('students.list_students'))
    errors, _ = data
    content = '\r\n'.join(errors)
    from io import BytesIO
    buf = BytesIO()
    buf.write(content.encode('utf-8-sig'))
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'导入错误日志_{key[:8]}.txt',
                     mimetype='text/plain; charset=utf-8')


@bp.route('/batch-edit-dormitory', methods=['POST'])
@perm_required('dormitory.manage')
def batch_edit_dormitory():
    """批量编辑宿舍相关信息（选科/住走读/出门权限/课本/班主任备注）"""
    ids = request.form.getlist('student_ids')
    if not ids:
        flash('未选择任何学生', 'warning')
        return redirect(url_for('students.list_students'))

    try:
        id_list = [int(i) for i in ids]
    except ValueError:
        flash('参数错误', 'danger')
        return redirect(url_for('students.list_students'))

    students = Student.query.filter(Student.id.in_(id_list)).all()
    if not students:
        flash('未找到选中的学生', 'warning')
        return redirect(url_for('students.list_students'))

    acc_fields = {
        'boarding_type': '住校/走读',
        'day_student_type': '出门权限',
        'textbook': '课本',
        'teacher_notes': '班主任备注',
    }
    student_fields = {
        'subject_selection': '选科',
    }

    updated_fields = []

    for field, label in acc_fields.items():
        val = request.form.get(field, '').strip()
        if val:
            for s in students:
                acc = StudentAccommodation.query.filter_by(student_id=s.id).first()
                if not acc:
                    acc = StudentAccommodation(student_id=s.id)
                    db.session.add(acc)
                setattr(acc, field, val)
            updated_fields.append(f'{label}={val}')

    for field, label in student_fields.items():
        val = request.form.get(field, '').strip()
        if val:
            for s in students:
                setattr(s, field, val)
            updated_fields.append(f'{label}={val}')

    if updated_fields:
        db.session.commit()
        log_operation(current_user, '批量修改', '学生', None,
                       f'{len(students)}名学生：{", ".join(updated_fields)}')
        flash(f'已更新{len(students)}名学生的：{", ".join(updated_fields)}', 'success')
    else:
        flash('未选择任何要修改的字段', 'warning')

    return redirect(url_for('students.list_students'))


@bp.route('/batch-delete', methods=['POST'])
@perm_required('students.import')
def batch_delete():
    """批量删除学生"""
    ids = request.form.getlist('student_ids')
    if not ids:
        flash('未选择任何学生', 'warning')
        return redirect(url_for('students.list_students'))

    try:
        id_list = [int(i) for i in ids]
    except ValueError:
        flash('参数错误', 'danger')
        return redirect(url_for('students.list_students'))

    students = Student.query.filter(Student.id.in_(id_list)).all()
    if not students:
        flash('未找到选中的学生', 'warning')
        return redirect(url_for('students.list_students'))

    count = len(students)
    for s in students:
        if s.bed_assignment:
            s.bed_assignment.student_id = None
        db.session.delete(s)
    db.session.commit()
    log_operation(current_user, '删除', '学生', None, f'批量删除 {count} 名学生')
    flash(f'已删除 {count} 名学生', 'success')
    return redirect(url_for('students.list_students'))


@bp.route('/<int:id>/transfer', methods=['POST'])
@perm_required('students.transfer')
def transfer(id):
    """单个学生调班调年级"""
    student = Student.query.get_or_404(id)
    # 权限控制：检查用户是否有权限调班该学生
    if current_user.role == 'homeroom_teacher':
        if student.grade != current_user.grade or student.class_name != current_user.class_name:
            flash('无权调班该学生', 'danger')
            return redirect(url_for('students.list_students'))
    elif current_user.role == 'grade_leader':
        if student.grade != current_user.grade:
            flash('无权调班该年级学生', 'danger')
            return redirect(url_for('students.list_students'))
    elif current_user.role not in ('admin',):
        flash('无权调班学生', 'danger')
        return redirect(url_for('students.list_students'))
    new_grade = request.form.get('new_grade', '').strip()
    new_class = request.form.get('new_class', '').strip()

    if not new_grade or not new_class:
        flash('年级和班级不能为空', 'danger')
        return redirect(url_for('students.list_students'))

    old_info = f'{student.grade} {student.class_name}'
    student.grade = new_grade
    student.class_name = new_class
    db.session.commit()
    log_operation(current_user, '更新', '学生', student.id, f'{student.name} {old_info}→{new_grade}{new_class}')
    flash(f'学生 {student.name} 已从 {old_info} 调至 {new_grade} {new_class}', 'success')
    return redirect(url_for('students.list_students'))


@bp.route('/download-transfer-template')
@perm_required('students.transfer')
def download_transfer_template():
    """下载调班模板（学号版或身份证号版）"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    tpl_type = request.args.get('type', 'student_number')

    wb = openpyxl.Workbook()
    ws = wb.active

    if tpl_type == 'id_card':
        ws.title = '身份证号调班模板'
        headers = ['身份证号', '新年级', '新班级']
        filename = '调班模板_身份证号版.xlsx'
        example_row = ['110101200801011234', '2025级', '02班']
        widths = [22, 12, 10]
    else:
        ws.title = '学号调班模板'
        headers = ['学号', '新年级', '新班级']
        filename = '调班模板_学号版.xlsx'
        example_row = ['20251001', '2025级', '02班']
        widths = [15, 12, 10]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    required_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = required_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.append(example_row)

    ws.cell(row=4, column=1, value='说明：').font = Font(bold=True, color='FF0000')
    ws.cell(row=5, column=1, value='1. 所有列均为必填项')
    if tpl_type == 'id_card':
        ws.cell(row=6, column=1, value='2. 身份证号必须与系统中已有学生匹配')
    else:
        ws.cell(row=6, column=1, value='2. 学号必须与系统中已有学生匹配')
    ws.cell(row=7, column=1, value='3. 新年级格式如：2025级')
    ws.cell(row=8, column=1, value='4. 新班级格式如：01班')
    ws.cell(row=9, column=1, value='5. 导入时请删除本说明和示例数据')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _get_changed_students(ws, match_col, grade_col, class_col, header_map, tpl_type):
    """从已处理的Excel中提取成功调班的学生列表"""
    from app.utils.crypto import encrypt as _enc
    students = []
    seen = set()
    for row_idx in range(2, ws.max_row + 1):
        match_val = ws.cell(row=row_idx, column=match_col).value
        if not match_val:
            continue
        match_val = str(match_val).strip()
        if not match_val or match_val in seen:
            continue
        seen.add(match_val)
        if tpl_type == 'id_card':
            student = Student.query.filter_by(_id_card_encrypted=_enc(match_val)).first()
        else:
            student = Student.query.filter_by(student_number=match_val).first()
        if student:
            students.append(student)
    return students


@bp.route('/download-class-transfer-template')
@perm_required('students.transfer')
def download_class_transfer_template():
    """下载批量调班模板（学号、姓名、当前班级、新班级）"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '批量调班'

    headers = ['学号', '姓名', '当前班级', '新年级', '新班级']

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    required_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    required_cols = {0, 3, 4}

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = required_fill if (col_idx - 1) in required_cols else header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    widths = [15, 10, 12, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.append(['20251001', '张三', '01班', '2025级', '02班'])

    ws.merge_cells('A2:E2')
    instr_cell = ws.cell(row=2, column=1,
        value='说明：橙色列必填 | 按学号匹配 | 第3行起填数据，删除本行和示例')
    instr_cell.font = Font(color='FF0000', bold=True, size=10)
    instr_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 30

    for col in range(1, len(headers) + 1):
        ws.cell(row=3, column=col).border = thin_border
        ws.cell(row=3, column=col).alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A1:E{ws.max_row}'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name='批量调班模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/batch-transfer', methods=['POST'])
@perm_required('students.transfer')
def batch_transfer():
    """批量调班"""
    file = request.files.get('file')
    tpl_type = request.form.get('tpl_type', 'student_number')

    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('请上传 .xlsx 格式的Excel文件', 'danger')
        return redirect(url_for('students.list_students'))

    try:
        import openpyxl.worksheet.datavalidation as dv
        _orig_init = dv.DataValidation.__init__
        def _patched_init(self, *a, **kw):
            kw.pop('id', None)
            _orig_init(self, *a, **kw)
        dv.DataValidation.__init__ = _patched_init

        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        header_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                header_map[str(val).strip()] = col

        # 确定匹配列
        if tpl_type == 'id_card':
            match_col = header_map.get('身份证号')
            match_label = '身份证号'
        else:
            match_col = header_map.get('学号')
            match_label = '学号'

        grade_col = header_map.get('新年级')
        class_col = header_map.get('新班级')

        if not match_col:
            flash(f'Excel缺少"{match_label}"列', 'danger')
            return redirect(url_for('students.list_students'))
        if not grade_col or not class_col:
            flash('Excel缺少"新年级"或"新班级"列', 'danger')
            return redirect(url_for('students.list_students'))

        errors = []
        updated_count = 0

        for row_idx in range(2, ws.max_row + 1):
            match_val = ws.cell(row=row_idx, column=match_col).value
            new_grade = ws.cell(row=row_idx, column=grade_col).value
            new_class = ws.cell(row=row_idx, column=class_col).value

            if match_val is None:
                continue

            match_val = str(match_val).strip()
            if not match_val:
                continue

            new_grade = str(new_grade).strip() if new_grade else ''
            new_class = str(new_class).strip() if new_class else ''

            row_label = f'第{row_idx}行'

            if not new_grade or not new_class:
                errors.append(f'{row_label}（{match_val}）：新年级或新班级为空')
                continue

            if tpl_type == 'id_card':
                student = Student.query.filter_by(_id_card_encrypted=_encrypt_id(match_val)).first()
            else:
                student = Student.query.filter_by(student_number=match_val).first()

            if not student:
                errors.append(f'{row_label}（{match_val}）：未找到该学生')
                continue

            student.grade = new_grade
            student.class_name = new_class
            updated_count += 1

        if errors:
            error_summary = f'调班完成，但有 {len(errors)} 条未处理：\n' + '\n'.join(errors[:20])
            if len(errors) > 20:
                error_summary += f'\n...还有 {len(errors) - 20} 条'
            if updated_count > 0:
                db.session.commit()
                # 写入变迁日志
                try:
                    changed = [{'id': s.id, 'student_number': s.student_number or '', 'name': s.name}
                              for s in _get_changed_students(ws, match_col, grade_col, class_col, header_map, tpl_type)]
                    if changed:
                        write_change_log('class', changed, detail='批量调班', operator_name=current_user.real_name)
                except Exception:
                    pass
                flash(f'成功调班 {updated_count} 名学生。{error_summary}', 'warning')
            else:
                db.session.rollback()
                flash(error_summary, 'danger')
        elif updated_count > 0:
            db.session.commit()
            # 写入变迁日志
            try:
                changed = [{'id': s.id, 'student_number': s.student_number or '', 'name': s.name}
                          for s in _get_changed_students(ws, match_col, grade_col, class_col, header_map, tpl_type)]
                if changed:
                    write_change_log('class', changed, detail='批量调班', operator_name=current_user.real_name)
            except Exception:
                pass
            flash(f'成功调班 {updated_count} 名学生', 'success')
        else:
            flash('Excel中没有有效的调班数据', 'warning')

    except Exception as e:
        db.session.rollback()
        flash(f'调班导入失败：{str(e)}', 'danger')

    return redirect(url_for('students.list_students'))


